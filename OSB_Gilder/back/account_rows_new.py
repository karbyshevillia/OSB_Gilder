import openpyxl as xl
import pandas as pd
import numpy as np
import re
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from indicators_new import IndicatorsFrameBuilder


class BalanceSheet:
    def __init__(self, parent_file, indicators_file, sheet):
        self.parent_file = parent_file
        self.indicators_file = indicators_file
        self.sheet = sheet

        # 1. High-speed parse to get coordinates
        # We use the 'grid' method from your original source for speed
        self.balance_codes_frame = self.parse_to_sane()
        print(self.balance_codes_frame.head(15))

        # 2. Initialize the builder and generate formulas
        self.builder = IndicatorsFrameBuilder(indicators_file, self.balance_codes_frame)
        self.indicators_frame = self.builder.build_frame()

    def get_col_letter(self, col_idx):
        """Converts a 0-based Pandas column index into an Excel column letter (0 -> 'A', 26 -> 'AA')"""
        string = ""
        col_idx += 1
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def parse_to_sane(self):
        """
        Fast extraction of row coordinates.
        Returns a DataFrame: ['Balance', 'Kind', 'Kind_Mark', 'Coord_Debit_Total', ...]
        """
        # 1. Fast Load with Calamine
        df_raw = pd.read_excel(self.parent_file, sheet_name=self.sheet.title, header=None, dtype=str, engine='calamine')

        # NEW: Lock in the original Excel Row Numbers before we do any slicing!
        # Pandas index is 0-based, Excel is 1-based.
        df_raw['Excel_Row'] = df_raw.index + 1

        # 2. Locate the Anchor
        anchor_mask = df_raw.apply(
            lambda row: row.astype(str).str.contains('Балансові рахунки', na=False, case=False).any(), axis=1)
        if anchor_mask.empty:
            raise ValueError("Could not find data anchor.")
        data_start_row = anchor_mask.idxmax()

        # ==========================================
        # 3. HEADER MERGED-CELL RESOLUTION & MAPPING
        # ==========================================
        headers = df_raw.iloc[:data_start_row].copy().astype(str).apply(lambda x: x.str.lower().str.strip())
        headers.replace({'nan': np.nan, 'none': np.nan, '': np.nan}, inplace=True)

        for idx, row in headers.iterrows():
            row_text_blob = " ".join(row.dropna().values)
            if any(kw in row_text_blob for kw in ['дебет', 'кредит', 'сальдо']):
                headers.loc[idx] = row.ffill()

        col_mapping = {}
        col_letters = {}  # NEW: Track the Excel Column letters!

        for col in df_raw.columns:
            if col == 'Excel_Row': continue
            text = " ".join(headers[col].dropna().tolist())

            mapped_name = None
            if 'дебет' in text:
                if 'усього' in text:
                    mapped_name = 'Debit_Total'
                elif 'нв' in text:
                    mapped_name = 'Debit_NC'
                elif 'ів' in text or 'iв' in text:
                    mapped_name = 'Debit_IC'
            elif 'кредит' in text:
                if 'усього' in text:
                    mapped_name = 'Credit_Total'
                elif 'нв' in text:
                    mapped_name = 'Credit_NC'
                elif 'ів' in text or 'iв' in text:
                    mapped_name = 'Credit_IC'
            elif 'сальдо' in text:
                if 'усього' in text:
                    mapped_name = 'Balance_Total'
                elif 'нв' in text:
                    mapped_name = 'Balance_NC'
                elif 'ів' in text or 'iв' in text:
                    mapped_name = 'Balance_IC'

            if mapped_name:
                col_mapping[col] = mapped_name
                col_letters[mapped_name] = self.get_col_letter(col)

        # ==========================================
        # 4. DATA BLOCK PROCESSING
        # ==========================================
        # Instead of resetting the index, we just copy. The 'Excel_Row' column travels with it safely.
        df_data = df_raw.iloc[data_start_row:].copy()

        # COLUMN 1: "Kind"
        df_data['Kind'] = pd.Series(np.nan, dtype='object', index=df_data.index)
        active_pattern = r'^\s*Актив(и|ні)?\s*$'
        passive_pattern = r'^\s*(Зобов\'язання|Пасив(и|ні)?|Капітал)\s*$'

        def is_exclusive_header_row(row, pattern):
            vals = [str(x).strip() for x in row if pd.notna(x) and str(x).strip().lower() not in ('', 'nan', 'none')]
            text_vals = [v for v in vals if not re.match(r'^-?[\d\s\.,]+$', v)]
            if not text_vals: return False
            return all(re.fullmatch(pattern, v, flags=re.IGNORECASE) for v in text_vals)

        # Note: drop 'Excel_Row' from the exclusivity check so it doesn't trigger false negatives!
        check_data = df_data.drop(columns=['Excel_Row'])
        active_mask = check_data.apply(lambda r: is_exclusive_header_row(r, active_pattern), axis=1)
        passive_mask = check_data.apply(lambda r: is_exclusive_header_row(r, passive_pattern), axis=1)

        df_data.loc[active_mask, 'Kind'] = 'Active'
        df_data.loc[passive_mask, 'Kind'] = 'Passive'
        df_data['Kind'] = df_data['Kind'].ffill()

        # COLUMNS 2-5: "Balance" via Purity Ratio & ffill
        best_balance_col = None
        max_purity = 0

        for col in df_data.columns:
            if col in col_mapping or col in ('Kind', 'Excel_Row'): continue

            sample = df_data[col].dropna().astype(str).str.replace(r'\.0$', '', regex=True)
            sample = sample[sample != 'nan']
            if sample.empty: continue

            four_digit_count = sample.str.match(r'^\d{4}$').sum()
            purity = four_digit_count / len(sample)

            if purity > max_purity and purity > 0.5:
                max_purity = purity
                best_balance_col = col

        if best_balance_col is None:
            raise ValueError("Could not find the Balance (Account Code) column.")

        col_mapping[best_balance_col] = 'Balance'
        col_letters['Balance'] = self.get_col_letter(best_balance_col)

        df_data[best_balance_col] = df_data[best_balance_col].astype(str).str.replace(r'\.0$', '', regex=True)
        df_data[best_balance_col] = df_data[best_balance_col].replace({'nan': np.nan, 'None': np.nan})
        df_data[best_balance_col] = df_data[best_balance_col].ffill()

        # COLUMN 7: "Kind_Mark" (А/П)
        ap_cols = []
        for col in df_data.columns:
            if col in col_mapping or col in ('Kind', 'Excel_Row'): continue
            sample = df_data[col].dropna().astype(str)
            if sample.str.match(r'^[АПA-P]{1,2}$', flags=re.IGNORECASE).sum() > len(sample) * 0.05:
                ap_cols.append(col)

        if not ap_cols:
            raise ValueError("Could not find the Kind_Mark (А/П) column.")
        col_mapping[ap_cols[0]] = 'Kind_Mark'
        col_letters['Kind_Mark'] = self.get_col_letter(ap_cols[0])

        # COLUMN 6: "Name"
        best_name_col = None
        max_name_len = 0
        for col in df_data.columns:
            if col not in col_mapping and col not in ('Kind', 'Excel_Row'):
                sample = df_data[col].dropna().astype(str)
                if not sample.empty:
                    avg_len = sample.str.len().mean()
                    if avg_len > max_name_len:
                        max_name_len = avg_len
                        best_name_col = col

        if best_name_col is not None:
            col_mapping[best_name_col] = 'Name'
            col_letters['Name'] = self.get_col_letter(best_name_col)

        df_data.rename(columns=col_mapping, inplace=True)

        if 'Name' in df_data.columns:
            df_data['Name'] = df_data['Name'].replace({'nan': np.nan, 'None': np.nan})
            df_data['Name'] = df_data['Name'].ffill()

        # ==========================================
        # 5. FILTERING AND ASSEMBLY
        # ==========================================

        line_items = df_data[
            df_data['Kind_Mark'].astype(str).str.match(r'^[АПA-P]{1,2}$', na=False, flags=re.IGNORECASE)].copy()

        line_items['Class'] = line_items['Balance'].str[0]
        line_items['Division'] = line_items['Balance'].str[:2]
        line_items['Group'] = line_items['Balance'].str[:3]

        kind_series = line_items['Kind'].copy()
        kind_series.loc[line_items['Class'] == '5'] = 'Capital'
        kind_series.loc[line_items['Class'] == '6'] = 'Revenue'
        kind_series.loc[line_items['Class'] == '7'] = 'Expenditures'

        invalid_class_mask = ~line_items['Class'].isin(['1', '2', '3', '4', '5', '6', '7', '9'])
        kind_series.loc[invalid_class_mask] = np.nan
        line_items['Kind'] = kind_series
        line_items['Kind'] = line_items['Kind'].replace({'nan': np.nan, 'None': np.nan})

        final_cols = [
            "Kind", "Class", "Division", "Group", "Balance", "Name", "Kind_Mark",
            "Debit_Total", "Debit_NC", "Debit_IC",
            "Credit_Total", "Credit_NC", "Credit_IC",
            "Balance_Total", "Balance_NC", "Balance_IC"
        ]

        for col in final_cols:
            if col not in line_items.columns:
                line_items[col] = np.nan

        # NEW: Inject Coordinates!
        # We will loop through columns that have an original physical location (like Balance_Total, Name, etc.)
        # and combine their Excel Letter + their Excel Row to create columns like 'Coord_Balance_Total'.
        coord_cols = []
        for col in final_cols:
            if col in col_letters:
                coord_col_name = f"Coord_{col}"
                coord_cols.append(coord_col_name)
                # Combine the tracked Letter (e.g., 'M') with the Row string (e.g., '15') -> 'M15'
                line_items[coord_col_name] = col_letters[col] + line_items['Excel_Row'].astype(str)

        # Assembly: Bring it all together (16 Data cols + Coordinate Cols)
        clean_df = line_items[final_cols + coord_cols].copy().dropna(axis=1, how="all")

        for col in final_cols[7:]:  # Safely cast the 9 value columns to floats
            clean_df[col] = pd.to_numeric(clean_df[col].astype(str).str.replace(',', ''), errors='coerce').fillna(0)

        return clean_df

    def insert_indicators_frame(self, start_coord="A1"):
        """
        Writes the indicators_frame with professional formatting and auto-column width.
        """
        # --- Define Styles ---
        header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        header_font = Font(bold=True, name='Calibri')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # --- Parse Coordinates ---
        match = re.match(r"([A-Z]+)(\d+)", start_coord)
        start_col_letter, start_row_num = match.groups()
        base_col = column_index_from_string(start_col_letter)
        base_row = int(start_row_num)

        # --- Write & Format Headers ---
        for c_idx, col_name in enumerate(self.indicators_frame.columns):
            cell = self.sheet.cell(row=base_row, column=base_col + c_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # --- Write & Format Data ---
        data = self.indicators_frame.values
        for r_idx, row_values in enumerate(data, start=1):
            for c_idx, value in enumerate(row_values):
                cell = self.sheet.cell(row=base_row + r_idx, column=base_col + c_idx)
                cell.value = value
                cell.border = thin_border

                # Names and IDs look better left-aligned, numbers/sums centered
                if c_idx < 2:  # ID and NAME columns
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

        # --- Auto-Adjust Column Widths ---
        # We estimate width based on the longest string in each column
        for i, col_name in enumerate(self.indicators_frame.columns):
            # Check header length and the max length in that column's data
            column_data = self.indicators_frame.iloc[:, i].astype(str)
            max_length = max(column_data.map(len).max(), len(col_name)) + 2

            # Hard cap at 50 so it doesn't get ridiculously wide for long formulas
            adjusted_width = min(max_length, 50)

            col_letter = get_column_letter(base_col + i)
            self.sheet.column_dimensions[col_letter].width = adjusted_width

if __name__ == '__main__':
    pd.set_option('display.max_columns', None)
    par_file = "/Users/illiaknu/Desktop/OSB_Gilder/OSB_Gilder/test_chamber/TEST_singular_9.xlsx"
    ind_file = "/Users/illiaknu/Desktop/OSB_Gilder/OSB_Gilder/back/testing/ind.csv"
    wb = xl.load_workbook(par_file)
    sheet = wb.worksheets[1]
    bs = BalanceSheet(par_file, ind_file, sheet)
    bs.insert_indicators_frame(start_coord="I1023")
    wb.save(par_file)