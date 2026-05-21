import openpyxl as xl
import pandas as pd
import numpy as np
import re
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .indicators_eval import IndicatorsFrameBuilder


class BalanceSheet:
    def __init__(self, parent_file, indicators_file, sheet, bank_class_xlsx=None):
        self.parent_file = parent_file
        self.indicators_file = indicators_file
        self.sheet = sheet
        self.start_cell_coord = f"G{sheet.max_row + 2}"
        self.bank_class_xlsx = bank_class_xlsx

        # 1. High-speed parse to get coordinates
        self.balance_codes_frame = self.parse_to_sane()

        # 2. Initialize the builder (We don't build_frame yet, we need the start_coord)
        self.builder = IndicatorsFrameBuilder(indicators_file, self.balance_codes_frame)
        self.indicators_frame = None

        # self.pivot_db = None
        # if bank_class_xlsx:
        #     self.create_db(bank_class_xlsx)

    def get_col_letter(self, col_idx):
        string = ""
        col_idx += 1
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def parse_to_sane(self):
        # 1. Fast Load with Calamine
        df_raw = pd.read_excel(self.parent_file, sheet_name=self.sheet.title, header=None, dtype=str, engine='calamine')

        df_raw['Excel_Row'] = df_raw.index + 1

        # 2. Locate the Anchor
        anchor_mask = df_raw.apply(
            lambda row: row.astype(str).str.contains('Балансові рахунки', na=False, case=False).any(), axis=1)
        if anchor_mask.empty:
            raise ValueError("Could not find data anchor.")
        data_start_row = anchor_mask.idxmax()

        # ==========================================
        # 3. HEADER MERGED-CELL RESOLUTION & MAPPING
        # ==========================================
        headers = df_raw.iloc[:data_start_row].copy().astype(str).apply(lambda x: x.str.lower().str.strip())
        headers.replace({'nan': np.nan, 'none': np.nan, '': np.nan}, inplace=True)

        for idx, row in headers.iterrows():
            row_text_blob = " ".join(row.dropna().values)
            if any(kw in row_text_blob for kw in ['дебет', 'кредит', 'сальдо']):
                headers.loc[idx] = row.ffill()

        col_mapping = {}
        col_letters = {}

        for col in df_raw.columns:
            if col == 'Excel_Row': continue
            text = " ".join(headers[col].dropna().tolist()).replace('i', 'і')  # Normalize i

            mapped_name = None

            # Find all occurrences of our primary keywords in the text
            primary_matches = re.findall(r'(дебет|кредит|сальдо)', text, re.IGNORECASE)

            if primary_matches:
                # Take the very last match found in the string (converted to lowercase for easy checking)
                last_primary = primary_matches[-1].lower()

                if last_primary == 'дебет':
                    if re.search(r'усього', text, re.IGNORECASE):
                        mapped_name = 'Debit_Total'
                    elif re.search(r'\bн\.?в\.?\b', text, re.IGNORECASE):
                        mapped_name = 'Debit_NC'
                    elif re.search(r'\bі\.?в\.?\b', text, re.IGNORECASE):
                        mapped_name = 'Debit_IC'

                elif last_primary == 'кредит':
                    if re.search(r'усього', text, re.IGNORECASE):
                        mapped_name = 'Credit_Total'
                    elif re.search(r'\bн\.?в\.?\b', text, re.IGNORECASE):
                        mapped_name = 'Credit_NC'
                    elif re.search(r'\bі\.?в\.?\b', text, re.IGNORECASE):
                        mapped_name = 'Credit_IC'

                elif last_primary == 'сальдо':
                    if re.search(r'усього', text, re.IGNORECASE):
                        mapped_name = 'Balance_Total'
                    elif re.search(r'\bн\.?в\.?\b', text, re.IGNORECASE):
                        mapped_name = 'Balance_NC'
                    elif re.search(r'\bі\.?в\.?\b', text, re.IGNORECASE):
                        mapped_name = 'Balance_IC'

            if mapped_name:
                col_mapping[col] = mapped_name
                col_letters[mapped_name] = self.get_col_letter(col)

        # ==========================================
        # 4. DATA BLOCK PROCESSING
        # ==========================================
        df_data = df_raw.iloc[data_start_row:].copy()

        df_data['Kind'] = pd.Series(np.nan, dtype='object', index=df_data.index)
        active_pattern = r'^\s*Актив(и|ні)?\s*$'
        passive_pattern = r'^\s*(Зобов\'язання|Пасив(и|ні)?|Капітал)\s*$'

        def is_exclusive_header_row(row, pattern):
            vals = [str(x).strip() for x in row if pd.notna(x) and str(x).strip().lower() not in ('', 'nan', 'none')]
            text_vals = [v for v in vals if not re.match(r'^-?[\d\s\.,]+$', v)]
            if not text_vals: return False
            return all(re.fullmatch(pattern, v, flags=re.IGNORECASE) for v in text_vals)

        check_data = df_data.drop(columns=['Excel_Row'])
        active_mask = check_data.apply(lambda r: is_exclusive_header_row(r, active_pattern), axis=1)
        passive_mask = check_data.apply(lambda r: is_exclusive_header_row(r, passive_pattern), axis=1)

        df_data.loc[active_mask, 'Kind'] = 'Active'
        df_data.loc[passive_mask, 'Kind'] = 'Passive'
        df_data['Kind'] = df_data['Kind'].ffill()

        best_balance_col = None
        max_purity = 0
        print(df_data.head())
        for col in df_data.columns:
            if col in col_mapping or col in ('Kind', 'Excel_Row'): continue

            sample = df_data[col].dropna().astype(str).str.replace(r'\.0$', '', regex=True)
            sample = sample[sample != 'nan']
            if sample.empty: continue

            four_digit_count = sample.str.match(r'^\d{4}$').sum()
            purity = four_digit_count / len(sample)

            if purity > max_purity and purity > 0.5:
                max_purity = purity
                best_balance_col = col

        if best_balance_col is None:
            raise ValueError("Could not find the Balance (Account Code) column.")

        col_mapping[best_balance_col] = 'Balance'
        col_letters['Balance'] = self.get_col_letter(best_balance_col)

        df_data[best_balance_col] = df_data[best_balance_col].astype(str).str.replace(r'\.0$', '', regex=True)
        df_data[best_balance_col] = df_data[best_balance_col].replace({'nan': np.nan, 'None': np.nan})
        df_data[best_balance_col] = df_data[best_balance_col].ffill()

        ap_cols = []
        for col in df_data.columns:
            if col in col_mapping or col in ('Kind', 'Excel_Row'): continue
            sample = df_data[col].dropna().astype(str)
            if sample.str.match(r'^[АПA-P]{1,2}$', flags=re.IGNORECASE).sum() > len(sample) * 0.05:
                ap_cols.append(col)

        if not ap_cols:
            raise ValueError("Could not find the Kind_Mark (А/П) column.")
        col_mapping[ap_cols[0]] = 'Kind_Mark'
        col_letters['Kind_Mark'] = self.get_col_letter(ap_cols[0])

        best_name_col = None
        max_name_len = 0
        for col in df_data.columns:
            if col not in col_mapping and col not in ('Kind', 'Excel_Row'):
                sample = df_data[col].dropna().astype(str)
                if not sample.empty:
                    avg_len = sample.str.len().mean()
                    if avg_len > max_name_len:
                        max_name_len = avg_len
                        best_name_col = col

        if best_name_col is not None:
            col_mapping[best_name_col] = 'Name'
            col_letters['Name'] = self.get_col_letter(best_name_col)

        df_data.rename(columns=col_mapping, inplace=True)

        if 'Name' in df_data.columns:
            df_data['Name'] = df_data['Name'].replace({'nan': np.nan, 'None': np.nan})
            df_data['Name'] = df_data['Name'].ffill()

        # ==========================================
        # 5. FILTERING AND ASSEMBLY
        # ==========================================
        line_items = df_data[
            df_data['Kind_Mark'].astype(str).str.match(r'^[АПA-P]{1,2}$', na=False, flags=re.IGNORECASE)].copy()

        line_items['Class'] = line_items['Balance'].str[0]
        line_items['Division'] = line_items['Balance'].str[:2]
        line_items['Group'] = line_items['Balance'].str[:3]

        kind_series = line_items['Kind'].copy()
        kind_series.loc[line_items['Class'] == '5'] = 'Capital'
        kind_series.loc[line_items['Class'] == '6'] = 'Revenue'
        kind_series.loc[line_items['Class'] == '7'] = 'Expenditures'

        invalid_class_mask = ~line_items['Class'].isin(['1', '2', '3', '4', '5', '6', '7', '9'])
        kind_series.loc[invalid_class_mask] = np.nan
        line_items['Kind'] = kind_series
        line_items['Kind'] = line_items['Kind'].replace({'nan': np.nan, 'None': np.nan})

        final_cols = [
            "Kind", "Class", "Division", "Group", "Balance", "Name", "Kind_Mark",
            "Debit_Total", "Debit_NC", "Debit_IC",
            "Credit_Total", "Credit_NC", "Credit_IC",
            "Balance_Total", "Balance_NC", "Balance_IC"
        ]

        for col in final_cols:
            if col not in line_items.columns:
                line_items[col] = np.nan

        coord_cols = []
        for col in final_cols:
            if col in col_letters:
                coord_col_name = f"Coord_{col}"
                coord_cols.append(coord_col_name)
                line_items[coord_col_name] = col_letters[col] + line_items['Excel_Row'].astype(str)

        clean_df = line_items[final_cols + coord_cols].copy().dropna(axis=1, how="all")

        for col in final_cols[7:]:
            clean_df[col] = pd.to_numeric(clean_df[col].astype(str).str.replace(',', ''), errors='coerce').fillna(0)

        return clean_df

    def insert_indicators_frame(self):
        # 1. Build the frame USING the target coordinate for proper IND_REF spatial pointers
        self.indicators_frame = self.builder.build_frame(self.start_cell_coord)

        match = re.match(r"([A-Z]+)(\d+)", self.start_cell_coord)
        start_col_letter, start_row_num = match.groups()
        base_col = column_index_from_string(start_col_letter)
        base_row = int(start_row_num)

        # 2. SAFETY UNMERGE: Prevent read-only crashes
        rows_to_write = len(self.indicators_frame) + 1
        cols_to_write = len(self.indicators_frame.columns)

        merged_ranges_to_unmerge = []
        for merged_range in list(self.sheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_range.bounds
            # Check for overlap between the target footprint and the merged range
            if (min_row <= base_row + rows_to_write and max_row >= base_row and
                    min_col <= base_col + cols_to_write and max_col >= base_col):
                merged_ranges_to_unmerge.append(str(merged_range))

        for m_range in merged_ranges_to_unmerge:
            self.sheet.unmerge_cells(m_range)

        # 3. Define Styles
        header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        header_font = Font(bold=True, name='Calibri')

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 4. Write & Format Headers
        for c_idx, col_name in enumerate(self.indicators_frame.columns):
            cell = self.sheet.cell(row=base_row, column=base_col + c_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 5. Write & Format Data
        data = self.indicators_frame.values
        for r_idx, row_values in enumerate(data, start=1):
            for c_idx, value in enumerate(row_values):
                cell = self.sheet.cell(row=base_row + r_idx, column=base_col + c_idx)
                cell.value = value
                cell.border = thin_border

                if c_idx < 2:  # ID and NAME
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

        # 6. Auto-Adjust Column Widths
        for i, col_name in enumerate(self.indicators_frame.columns):
            column_data = self.indicators_frame.iloc[:, i].astype(str)
            max_length = max(column_data.map(len).max(), len(col_name)) + 2
            adjusted_width = min(max_length, 50)
            col_letter = get_column_letter(base_col + i)
            self.sheet.column_dimensions[col_letter].width = adjusted_width

    @property
    def create_db(self):
        """
        Creates a flat database frame from the parsed balance codes,
        appending bank identification from the sheet title and the nested Excel classification file.
        Duplicates rows for multiple indicators using .explode().
        """
        # 1. Parse Bank Code and Name from Sheet Title
        match = re.match(r'^\s*(\d+)', self.sheet.title)
        bank_code = match.group(1) if match else ""
        bank_name = self.sheet.title[match.end():].strip() if match else self.sheet.title

        classification_xlsx_path = self.bank_class_xlsx
        # 2. Determine Bank Class from Excel (.xlsx)
        bank_class = "інше"  # Default fallback matching the new structural labels
        if classification_xlsx_path:
            try:
                # Load classification spreadsheet via Pandas
                class_df = pd.read_excel(classification_xlsx_path)

                # Normalize column headers to lowercase & stripped strings to avoid syntax mismatches
                class_df.columns = class_df.columns.str.lower().str.strip()

                if 'nkb' in class_df.columns and 'class' in class_df.columns:
                    matched_row = None

                    # LAYER 1: Primary lookup matching exact NKB code (Bulletproof)
                    if bank_code:
                        try:
                            target_nkb = int(bank_code)
                            class_df['nkb_numeric'] = pd.to_numeric(class_df['nkb'], errors='coerce')
                            match_mask = class_df['nkb_numeric'] == target_nkb
                            if match_mask.any():
                                matched_row = class_df[match_mask].iloc[0]
                        except Exception:
                            pass

                    # LAYER 2: Secondary text-fallback if NKB code wasn't matched or doesn't exist
                    if matched_row is None and 'name' in class_df.columns:
                        def clean_for_match(name):
                            if not name: return ""
                            s = re.sub(r'[\s\-\,\.\(\)\'\"«»指標“”`]', '', str(name).lower())
                            return s.translate(str.maketrans('ііїєсхрАТМВНЕО', 'iiiеcxpATMBHEO'))

                        b_name_clean = clean_for_match(bank_name)

                        for _, row in class_df.iterrows():
                            xlsx_b_name = clean_for_match(row.get("name", ""))
                            if b_name_clean and xlsx_b_name and (
                                    b_name_clean in xlsx_b_name or xlsx_b_name in b_name_clean):
                                matched_row = row
                                break

                    # Extract the string value from the custom matrix if found
                    if matched_row is not None:
                        bank_class = str(matched_row['class']).strip()

            except Exception as e:
                print(f"Warning: Could not load classification Excel ({e}).")

        # 3. Build the Database Frame
        db_df = self.balance_codes_frame.copy()
        translation = {"Active": "Активи",
                       "Passive": "Пасиви",
                       "Capital": "Капітал",
                       "Revenue": "Доходи",
                       "Expenditures": "Витрати"}
        for k, v in translation.items():
            db_df.loc[db_df["Kind"] == k, "Kind"] = v

        # --- THE EXPLODE LOGIC FOR MANY-TO-MANY INDICATORS ---
        def get_indicators_for_row(idx):
            print(self.builder.code_usage)
            df = self.builder.rules
            pd.set_option('display.max_columns', None)
            print(df.head(3))
            used_by = self.builder.code_usage.get(idx, set())
            print(used_by)
            if not used_by:
                return ["(без привʼязки)"]
            # return [str(i_id) for i_id in sorted(used_by)]
            print([df.loc[df["ID"] == i_id, "NAME"] for i_id in sorted(used_by)])
            return [str(df.at[int(i_id) - 1, "NAME"]) for i_id in sorted(used_by)] + ["(без привʼязки)"]

        # Assign lists to the column and explode
        db_df['indicator'] = db_df.index.map(get_indicators_for_row)
        db_df = db_df.explode('indicator', ignore_index=True)
        # -----------------------------------------------------

        # Insert metadata columns at the front
        db_df.insert(0, "bank_class", bank_class)
        db_df.insert(0, "bank_name", bank_name)
        db_df.insert(0, "bank_code", bank_code)

        # Drop the Excel-specific coordinate columns as they aren't needed in a DB
        coord_cols = [c for c in db_df.columns if str(c).startswith('Coord_') or c == 'Excel_Row']
        db_df.drop(columns=coord_cols, inplace=True, errors='ignore')

        # Move indicator to the desired column position (index 4)
        indicator_col = db_df.pop("indicator")
        db_df.insert(4, "indicator", indicator_col)

        self.pivot_db = db_df
        return db_df


if __name__ == '__main__':
    pd.set_option('display.max_columns', None)

    par_file = "/Users/illiaknu/Desktop/OSB_Gilder/OSB_Gilder/test_chamber/DEBUG_2024_2.xlsx"
    ind_file = "/Users/illiaknu/Desktop/OSB_Gilder/OSB_Gilder/back/testing/ind.csv"
    banks_file = "/Users/illiaknu/Desktop/OSB_Gilder/OSB_Gilder/back/testing/classification.xlsx"

    wb = xl.load_workbook(par_file)
    sheet = wb.worksheets[0]

    bs = BalanceSheet(par_file, ind_file, sheet, bank_class_xlsx=banks_file)

    # Try a notoriously problematic coordinate to test the unmerge feature
    bs.insert_indicators_frame()

    wb.save(par_file)
    print(bs.pivot_db.head(15))
    print("Injection Complete!")