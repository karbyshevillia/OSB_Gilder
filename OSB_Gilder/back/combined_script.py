"""
Combined Bank Data Processing Script


This script combines the functionality of:
1. create_sheet_index.py - Creates and formats the Index sheet
2. Danya's_code.py - Fills the Index sheet with formulas from bank sheets

Process:
1. Create Index sheet with headers and formatting
2. Add hyperlinks to all bank sheets
3. Fill formulas linking to indicators in bank sheets
4. Calculate column sums
5. Save the file
"""

import openpyxl as xl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
# from openpyxl.xml.functions import fromstring
from xml.etree.ElementTree import fromstring
import sys
import io
import re
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from .account_rows import BalanceSheet

# Ensure UTF-8 encoding for Cyrillic characters
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
if sys.stderr.encoding != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


def _datalabel_text_style(size_pt=12, bold=True, color="000000"):
    size_100 = int(size_pt * 100)
    bold_flag = "1" if bold else "0"
    xml = f'''
<a:txPr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
    <a:bodyPr/>
    <a:lstStyle/>
    <a:p>
        <a:pPr>
            <a:defRPr sz="{size_100}" b="{bold_flag}">
                <a:solidFill><a:srgbClr val="{color}"/></a:solidFill>
            </a:defRPr>
        </a:pPr>
        <a:endParaRPr lang="uk-UA"/>
    </a:p>
</a:txPr>
'''
    return RichText.from_tree(fromstring(xml))


def make_formula(sheet_name, column, row):
    """Create Excel formula reference to another sheet."""
    return f"='{sheet_name}'!${column}${row}"


def safe_set_cell_value(sheet, row, col, value):
    """Safely set cell value, handling merged cells"""
    target_cell = sheet.cell(row=row, column=col)

    # Check if it's a merged cell
    if isinstance(target_cell, MergedCell):
        # Find the merged range and unmerge it temporarily
        for merged_range in list(sheet.merged_cells.ranges):
            if target_cell.coordinate in merged_range:
                sheet.unmerge_cells(str(merged_range))
                # Now write to the cell
                sheet.cell(row=row, column=col).value = value
                return

    # If not merged, just set the value normally
    target_cell.value = value


def code_bank(name):
    """Extract bank code from sheet name."""
    match = re.match(r"\s*(\d+)", name)
    return (match.group(1), str(name)) if match else None


def is_valid_sheet_name(sheet_name):
    """Check if a sheet name matches the required pattern (starts with a number)."""
    pattern = r'^\s*\d+'
    return re.match(pattern, sheet_name) is not None


def create_index_sheet(workbook):
    """
    Create or replace the Index sheet.

    Args:
        workbook: openpyxl Workbook object

    Returns:
        Worksheet object for the Index sheet
    """
    # Check if Index sheet already exists
    if "Index" in workbook.sheetnames:
        print("\n'Index' sheet already exists - removing it...")
        index_sheet = workbook["Index"]
        workbook.remove(index_sheet)
        print("  ✓ Existing Index sheet removed")

    # Create new Index sheet at the beginning
    print("Creating new Index sheet...")
    index_sheet = workbook.create_sheet("Index", 0)
    print("  ✓ Index sheet created")

    return index_sheet

# def create_db_sheet(sheets_dict, workbook):
#     ws = workbook.create_sheet("Combined_Data")
#     df = pd.concat([bs.sheet_db for bs in sheets_dict.values()])
#     for row in dataframe_to_rows(df, index=False, header=True):
#         ws.append(row)
#     tab = Table(displayName="Combined_Data", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
#
#     # 3. Add a visual style (optional but recommended)
#     style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
#                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
#     tab.tableStyleInfo = style
#
#     # 4. Add the table to the worksheet
#     ws.add_table(tab)

def create_pivot_sheet(sheets_dict, workbook):
    ws = workbook.create_sheet("Pivot_Data")
    df = pd.concat([bs.pivot_db for bs in sheets_dict.values()])
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    tab = Table(displayName="Pivot_Data", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")

    # 3. Add a visual style (optional but recommended)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    # 4. Add the table to the worksheet
    ws.add_table(tab)


def format_header(sheet):
    """
    Format the header rows of the Index sheet.
    """
    # Hide description rows
    for row_id in [1, 2, 3, 4]:
        sheet.row_dimensions[row_id].hidden = True

    # Description rows
    sheet['C2'] = "Оборотно-сальдовий баланс банку*"
    sheet['D2'] = "відповідно до Додатку 1 до постанови Правління Національного банку України"
    sheet[
        'D3'] = "від 15 березня 2018 року № 11 \"Про встановлення переліку інформації, що підлягає обов'язковому опублікуванню банками України\" (зі змінами)"
    sheet['D4'] = "за даними статистичної звітності з файлів 02X \"Дані про обороти та залишки на рахунках\"."

    # Row 5: Column headers
    sheet['A5'] = "NKB"
    sheet['B5'] = "Порядковий номер"
    sheet['C5'] = "Назва банку"
    sheet['D5'] = "МФО"

    header_fill = PatternFill(start_color="0070C0",
                              end_color="0070C0",
                              fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format A-D columns
    for col in ['A5', 'B5', 'C5', 'D5']:
        sheet[col].font = Font(name='Arial', size=11, bold=True, color="000000")
        sheet[col].alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        sheet[col].border = thin_border

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 15

    # E-AD headers
    headers_e_ad = {
        5: "ПРОЦЕНТИ ЗА ДЕПОЗИТНИМИ СЕРТИФІКАТАМИ (ДС), тис. грн",
        6: "ПРОЦЕНТИ ЗА ОВДП (ОВДП), тис.грн",
        7: "ПРОЦЕНТИ  ЗА КРЕДИТАМИ ЗДУ",
        8: "ПРОЦЕНТИ ЗА КРЕДИТАМИ БІЗНЕСУ",
        9: "ПРОЦЕНТИ ЗА КРЕДИТАМИ НАСЕЛЕННЮ",
        10: "ПРОЦЕНТНІ ДОХОДИ ВСЬОГО (ПД), тис. грн",
        11: "ДС/ПД, %",
        12: "ДС до підсумку, %",
        13: "(ДС+ОВДП)/ПД, %",
        14: "ОВДП/ПД, %",
        15: "ОВДП до підсумку",
        16: "Обсяг ДС дебит (середні, поділено на кількість банк.днів)",
        17: "Обсяг ДС кредит (середні, поділено на кількість банк.днів)",
        18: "Кошти в НБУ (кор.рахунок) (кредит) (код 1200)",
        19:"Кошти в НБУ (УСЬОГО) (кредит) (розділ 12)",
        20:"Проценти за поточними вкладами ДГ (код 7040)",
        21:"Проценти за строковими вкладами ДГ (код 7041)",
        22:"Кредити субʼєктам господарювання (р.20, р.23, 260)",
        23:"Кредити ЗДУ (р.21)",
        24:"Кредити фізичним особам (р. 22, 24, 262)",
        25:"Кредити небанківським установам (265)",
        26:"Прибуток звітного року (5040)",
        27:"Збиток звітного року (5041)",
        28:"Податок на прибуток (7900)",
        29:"Інші податки (741)",
        30:"Відрахування в резерви (770)",
        31:"Загальні доходи (6)",
        32:"Загальні витрати (7)"
    }

    # Blue headers (columns E-J / 5-10)
    for col_idx in range(5, 11):
        col_letter = get_column_letter(col_idx)
        cell = sheet[f'{col_letter}5']
        cell.value = headers_e_ad.get(col_idx, "")
        cell.font = Font(name='Arial', size=9, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        sheet.column_dimensions[col_letter].width = 18
        sheet.row_dimensions[5].height = 60

    # Regular headers (columns K-AF / 11-32)
    for col_idx in range(11, 40):
        col_letter = get_column_letter(col_idx)
        cell = sheet[f'{col_letter}5']
        cell.value = headers_e_ad.get(col_idx, "")
        cell.font = Font(name='Arial', size=9, bold=True, color="000000")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        sheet.column_dimensions[col_letter].width = 18
        sheet.row_dimensions[5].height = 60

    # Freeze first 4 columns and header row 5
    sheet.freeze_panes = "E6"

    print("  ✓ Headers formatted")


def add_hyperlinks(workbook, index_sheet, matching_sheets):
    """
    Add hyperlinks to matching sheets starting from row 6.

    Args:
        workbook: openpyxl Workbook object
        index_sheet: The Index worksheet
        matching_sheets: List of sheet names that match the pattern
    """
    print("\nAdding hyperlinks...")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 6
    index_counter = 1

    for sheet_name in matching_sheets:
        match = re.match(r'^\s*(\d+)', sheet_name)
        number_str = match.group(1) if match else ""
        text_part = sheet_name[match.end():].strip() if match else sheet_name

        # Column A: Number with hyperlink
        cell_a = index_sheet.cell(row=current_row, column=1)
        cell_a.value = number_str
        cell_a.hyperlink = f"#'{sheet_name}'!A1"
        cell_a.font = Font(name='Arial', size=11, color="0000FF", underline="single")
        cell_a.border = thin_border
        cell_a.alignment = Alignment(horizontal='center', vertical='center')

        # Column B: Index (1, 2, 3, ...)
        cell_b = index_sheet.cell(row=current_row, column=2)
        cell_b.value = index_counter
        cell_b.font = Font(name='Arial', size=11)
        cell_b.border = thin_border
        cell_b.alignment = Alignment(horizontal='center', vertical='center')

        # Column C: Sheet name without number with hyperlink
        cell_c = index_sheet.cell(row=current_row, column=3)
        cell_c.value = text_part
        cell_c.hyperlink = f"#'{sheet_name}'!A1"
        cell_c.font = Font(name='Arial', size=11, color="0000FF", underline="single")
        cell_c.border = thin_border
        cell_c.alignment = Alignment(horizontal='left', vertical='center')

        # Column D: MFO (if found in "Зміст" sheet)
        mfo_value = None
        if "Зміст" in workbook.sheetnames and number_str:
            contents = workbook["Зміст"]
            for row_idx in range(1, contents.max_row + 1):
                cell_val = contents.cell(row=row_idx, column=1).value
                if cell_val is None:
                    continue
                try:
                    cell_text = str(cell_val).strip()
                    if cell_text == number_str or (cell_text.isdigit() and int(cell_text) == int(number_str)):
                        mfo_value = contents.cell(row=row_idx, column=3).value
                        break
                except Exception:
                    pass

        cell_d = index_sheet.cell(row=current_row, column=4)
        if mfo_value is not None:
            cell_d.value = str(mfo_value).strip()
            cell_d.hyperlink = f"#'{sheet_name}'!A1"
            cell_d.font = Font(name='Arial', size=11, color="0000FF", underline="single")
        cell_d.border = thin_border
        cell_d.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1
        index_counter += 1

    # Adjust row height for better readability
    for row in range(6, current_row):
        index_sheet.row_dimensions[row].height = 22

    print(f"  ✓ Added {len(matching_sheets)} hyperlinks")


def create_sheets_dict(workbook, parent_file):
    """Create dictionary of valid bank sheets with BalanceSheet objects."""
    print("\n=== STEP 2: Creating BalanceSheet objects ===")
    sheets_dict = {}

    sheet_num = 0
    total_sheets = len(workbook.worksheets)

    for sheet in workbook.worksheets:
        sheet_num += 1
        title = code_bank(sheet.title)
        if title != None:
            print(f"  [{sheet_num}/{total_sheets}] Processing sheet: {sheet.title}")
            sheet_value = BalanceSheet(sheet=sheet, parent_file=parent_file)
            sheets_dict[title] = sheet_value
            print(f"      ✓ Added to dictionary")
        else:
            print(f"  [{sheet_num}/{total_sheets}] Skipping: {sheet.title} (no bank code)")

    print(f"\n✓ Created dictionary with {len(sheets_dict)} bank sheets\n")
    return sheets_dict


def fill_formulas(workbook, vidpovidnist_pokaznykiv, parent_file):
    """
    Fill Index sheet with formulas linking to bank sheets.

    This is the core logic from Danya's_code.py
    """
    content_sheet = workbook["Index"]

    sheets_dict = create_sheets_dict(workbook, parent_file)
    vp = vidpovidnist_pokaznykiv
    print(f"  ✓ Dictionary has {len(vp)} indicators to match")

    sheet_count = 0
    total_sheets = len(sheets_dict)
    print(f"\n=== STEP 4: Processing {total_sheets} bank sheets ===")

    for codes in sheets_dict:
        sheet_count += 1
        obyect = sheets_dict[codes]
        code = codes[0]
        name_of_sheet = codes[1]

        matrix_base = obyect.indicator_frame_cells
        pokaznyk_row = 5
        index_column = 1
        started_row = 6
        print(f"\n[{sheet_count}/{total_sheets}] Processing: {codes}")
        print(f"  → Finding bank code {code} in Index sheet...")

        # Find the row in Index sheet that corresponds to this bank
        for row in content_sheet.iter_rows(
                min_row=started_row,
                max_row=content_sheet.max_row,
                min_col=index_column,
                max_col=index_column):
            cell = row[0]
            if cell.value != None and int(cell.value) == int(code):
                main_bank_row = cell.row
                print(f"  ✓ Found at row {main_bank_row}")
                break

        print(f"  → Matching indicators from matrix...")

        # For each indicator in the bank's data
        for row in matrix_base.itertuples():
            pokaznyk = row.Index

            # Find matching column header in Index sheet
            for column in content_sheet.iter_cols(min_row=pokaznyk_row, max_row=pokaznyk_row, min_col=5, max_col=100):
                cell = column[0]

                if vp.get(str(cell.value)) != None and vp.get(str(cell.value))[0] == str(pokaznyk):
                    print(f"     ✓ Match: {str(cell.value)}")
                    hnp = vp[str(cell.value)][1]
                    cel = row[hnp]
                    row_link = cel.row
                    column_link = cel.column_letter
                    cell_link = content_sheet.cell(row=main_bank_row, column=cell.column)
                    cell_link.value = make_formula(name_of_sheet, column_link, row_link)
                    break

        print(f"  ✓ Completed sheet {sheet_count}/{total_sheets}")

    print(f"\n✓ All bank sheets processed!")

    # create_db_sheet(sheets_dict, workbook)
    create_pivot_sheet(sheets_dict, workbook)


def sum_columns_by_header(workbook, workbook_values, sheet, data_last_row):
    """
    Write totals row under data using Excel formulas.
    Group rows are also formula-based and only filled for columns
    where row 5 header is not empty.
    """
    results = {}
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    def _to_float(value):
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.replace(' ', '').replace(',', '.')
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        return 0.0

    def _resolve_value(value, depth=0):
        if depth > 2:
            return 0.0

        if isinstance(value, str):
            # Handle links like ='SHEET'!$C$10
            match = re.match(r"^='(.+)'!\$?([A-Z]+)\$?(\d+)$", value)
            if match:
                src_sheet_name = match.group(1)
                src_col = match.group(2)
                src_row = int(match.group(3))

                if src_sheet_name in workbook.sheetnames:
                    src_ws = workbook[src_sheet_name]
                    src_value = src_ws[f"{src_col}{src_row}"].value
                    # Prefer cached calculated value (data_only workbook)
                    if src_sheet_name in workbook_values.sheetnames:
                        src_ws_values = workbook_values[src_sheet_name]
                        cached_value = src_ws_values[f"{src_col}{src_row}"].value
                        if cached_value is not None:
                            return _to_float(cached_value)

                    return _resolve_value(src_value, depth + 1)
                return 0.0

        return _to_float(value)

    print("\n=== STEP 5: Calculating column sums ===")

    # Write totals row under all bank rows
    total_row = data_last_row + 1
    sheet.cell(row=total_row, column=1).value = "Всього"
    sheet.cell(row=total_row, column=1).font = Font(name='Arial', size=10, bold=True)
    sheet.cell(row=total_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # Keep list of columns where row-5 header is not empty
    non_empty_header_cols = []

    # Map row-5 headers to column indexes
    header_to_col = {}
    for col_idx in range(5, sheet.max_column + 1):
        header_value = sheet.cell(row=5, column=col_idx).value
        if header_value is not None and str(header_value).strip() != "":
            header_to_col[str(header_value)] = col_idx

    ds_col_idx = header_to_col.get("ПРОЦЕНТИ ЗА ДЕПОЗИТНИМИ СЕРТИФІКАТАМИ (ДС), тис. грн")
    ovdp_col_idx = header_to_col.get("ПРОЦЕНТИ ЗА ОВДП (ОВДП), тис.грн")
    pd_col_idx = header_to_col.get("ПРОЦЕНТНІ ДОХОДИ ВСЬОГО (ПД), тис. грн")
    ds_pd_col_idx = header_to_col.get("ДС/ПД, %")
    ds_share_col_idx = header_to_col.get("ДС до підсумку, %")
    ds_ovdp_pd_col_idx = header_to_col.get("(ДС+ОВДП)/ПД, %")
    ovdp_pd_col_idx = header_to_col.get("ОВДП/ПД, %")
    ovdp_share_col_idx = header_to_col.get("ОВДП до підсумку")
    finrez_col_idx = header_to_col.get("ФінРез (ОВДП-ДС)")

    # Fill DS/PD formula for each bank row (same-row calculation)
    if ds_col_idx is not None and pd_col_idx is not None and ds_pd_col_idx is not None:
        ds_col_letter = get_column_letter(ds_col_idx)
        pd_col_letter = get_column_letter(pd_col_idx)
        ds_pd_col_letter = get_column_letter(ds_pd_col_idx)

        for row_idx in range(6, data_last_row + 1):
            ds_pd_cell = sheet[f"{ds_pd_col_letter}{row_idx}"]
            ds_pd_cell.value = f"=IFERROR({ds_col_letter}{row_idx}/{pd_col_letter}{row_idx}*100,0)"
            ds_pd_cell.number_format = '0.0'

    # Fill DS share formula for each bank row (row DS / total DS column)
    if ds_col_idx is not None and ds_share_col_idx is not None:
        ds_col_letter = get_column_letter(ds_col_idx)
        ds_share_col_letter = get_column_letter(ds_share_col_idx)

        for row_idx in range(6, data_last_row + 1):
            ds_share_cell = sheet[f"{ds_share_col_letter}{row_idx}"]
            ds_share_cell.value = f"=IFERROR({ds_col_letter}{row_idx}/SUM({ds_col_letter}6:{ds_col_letter}{data_last_row})*100,0)"
            ds_share_cell.number_format = '0.0'

    # Fill (DS+OVDP)/PD formula for each bank row
    if ds_col_idx is not None and ovdp_col_idx is not None and pd_col_idx is not None and ds_ovdp_pd_col_idx is not None:
        ds_col_letter = get_column_letter(ds_col_idx)
        ovdp_col_letter = get_column_letter(ovdp_col_idx)
        pd_col_letter = get_column_letter(pd_col_idx)
        ds_ovdp_pd_col_letter = get_column_letter(ds_ovdp_pd_col_idx)

        for row_idx in range(6, data_last_row + 1):
            ds_ovdp_pd_cell = sheet[f"{ds_ovdp_pd_col_letter}{row_idx}"]
            ds_ovdp_pd_cell.value = f"=IFERROR(({ds_col_letter}{row_idx}+{ovdp_col_letter}{row_idx})/{pd_col_letter}{row_idx}*100,0)"
            ds_ovdp_pd_cell.number_format = '0.0'

    # Fill OVDP/PD formula for each bank row
    if ovdp_col_idx is not None and pd_col_idx is not None and ovdp_pd_col_idx is not None:
        ovdp_col_letter = get_column_letter(ovdp_col_idx)
        pd_col_letter = get_column_letter(pd_col_idx)
        ovdp_pd_col_letter = get_column_letter(ovdp_pd_col_idx)

        for row_idx in range(6, data_last_row + 1):
            ovdp_pd_cell = sheet[f"{ovdp_pd_col_letter}{row_idx}"]
            ovdp_pd_cell.value = f"=IFERROR({ovdp_col_letter}{row_idx}/{pd_col_letter}{row_idx}*100,0)"
            ovdp_pd_cell.number_format = '0.0'

    # Fill OVDP share formula for each bank row (row OVDP / total OVDP column)
    if ovdp_col_idx is not None and ovdp_share_col_idx is not None:
        ovdp_col_letter = get_column_letter(ovdp_col_idx)
        ovdp_share_col_letter = get_column_letter(ovdp_share_col_idx)

        for row_idx in range(6, data_last_row + 1):
            ovdp_share_cell = sheet[f"{ovdp_share_col_letter}{row_idx}"]
            ovdp_share_cell.value = f"=IFERROR({ovdp_col_letter}{row_idx}/SUM({ovdp_col_letter}6:{ovdp_col_letter}{data_last_row})*100,0)"
            ovdp_share_cell.number_format = '0.0'

    # Fill FinRez formula for each bank row: DS + OVDP
    if ds_col_idx is not None and ovdp_col_idx is not None and finrez_col_idx is not None:
        ds_col_letter = get_column_letter(ds_col_idx)
        ovdp_col_letter = get_column_letter(ovdp_col_idx)
        finrez_col_letter = get_column_letter(finrez_col_idx)

        for row_idx in range(6, data_last_row + 1):
            finrez_cell = sheet[f"{finrez_col_letter}{row_idx}"]
            finrez_cell.value = f"=IFERROR({ds_col_letter}{row_idx}+{ovdp_col_letter}{row_idx},0)"
            finrez_cell.number_format = '#,##0.00'

    # Iterate through data columns (E and onward)
    for col_idx in range(5, sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header_cell = sheet[f'{col_letter}5']
        header_value = header_cell.value

        if header_value is None or str(header_value).strip() == "":
            continue

        non_empty_header_cols.append(col_idx)
        sum_cell = sheet[f"{col_letter}{total_row}"]

        # Special handling for percentage columns
        if ds_pd_col_idx is not None and col_idx == ds_pd_col_idx:
            if ds_col_idx is not None and pd_col_idx is not None:
                ds_col_letter = get_column_letter(ds_col_idx)
                pd_col_letter = get_column_letter(pd_col_idx)
                sum_cell.value = f"=IFERROR({ds_col_letter}{total_row}/{pd_col_letter}{total_row}*100,0)"
            else:
                sum_cell.value = "=0"
        elif ds_share_col_idx is not None and col_idx == ds_share_col_idx:
            if ds_col_idx is not None:
                ds_col_letter = get_column_letter(ds_col_idx)
                sum_cell.value = f"=IFERROR({ds_col_letter}{total_row}/SUM({ds_col_letter}6:{ds_col_letter}{data_last_row})*100,0)"
            else:
                sum_cell.value = "=0"
        elif ds_ovdp_pd_col_idx is not None and col_idx == ds_ovdp_pd_col_idx:
            if ds_col_idx is not None and ovdp_col_idx is not None and pd_col_idx is not None:
                ds_col_letter = get_column_letter(ds_col_idx)
                ovdp_col_letter = get_column_letter(ovdp_col_idx)
                pd_col_letter = get_column_letter(pd_col_idx)
                sum_cell.value = f"=IFERROR(({ds_col_letter}{total_row}+{ovdp_col_letter}{total_row})/{pd_col_letter}{total_row}*100,0)"
            else:
                sum_cell.value = "=0"
        elif ovdp_pd_col_idx is not None and col_idx == ovdp_pd_col_idx:
            if ovdp_col_idx is not None and pd_col_idx is not None:
                ovdp_col_letter = get_column_letter(ovdp_col_idx)
                pd_col_letter = get_column_letter(pd_col_idx)
                sum_cell.value = f"=IFERROR({ovdp_col_letter}{total_row}/{pd_col_letter}{total_row}*100,0)"
            else:
                sum_cell.value = "=0"
        elif ovdp_share_col_idx is not None and col_idx == ovdp_share_col_idx:
            if ovdp_col_idx is not None:
                ovdp_col_letter = get_column_letter(ovdp_col_idx)
                sum_cell.value = f"=IFERROR({ovdp_col_letter}{total_row}/SUM({ovdp_col_letter}6:{ovdp_col_letter}{data_last_row})*100,0)"
            else:
                sum_cell.value = "=0"
        else:
            sum_cell.value = f"=IFERROR(SUM({col_letter}6:{col_letter}{data_last_row}),0)"

        sum_cell.font = Font(name='Arial', size=10, bold=True)
        sum_cell.alignment = Alignment(horizontal='center', vertical='center')
        sum_cell.number_format = '#,##0.00'

        results[str(header_value)] = {
            'column': col_letter,
            'formula': sum_cell.value
        }
        print(f"  ✓ Column {col_letter}: {header_value}")
        print(f"    Formula: {sum_cell.value}")

    if not non_empty_header_cols:
        print("  ⚠ No non-empty headers found from row 5")
        return results

    last_formula_col = max(non_empty_header_cols)

    # Apply row styling (borders/fill) across entire Index width
    for col_idx in range(1, last_formula_col + 1):
        cell = sheet.cell(row=total_row, column=col_idx)
        cell.border = thin_border
        cell.fill = total_fill

    sheet.row_dimensions[total_row].height = 22

    # Category totals by fixed bank groups
    def _normalize_name(name):
        if name is None:
            return ""
        return re.sub(r"\s+", " ", str(name)).strip()

    state_banks = {
        _normalize_name('АТ "Укрексімбанк"'),
        _normalize_name('АТ "ОЩАДБАНК"'),
        _normalize_name('АТ КБ "ПриватБанк"'),
        _normalize_name('АТ "СЕНС БАНК"'),
        _normalize_name('АБ "УКРГАЗБАНК"'),
        _normalize_name('АТ "ПЕРШИЙ ІНВЕСТИЦІЙНИЙ БАНК"'),
        _normalize_name('АТ "МОТОР-БАНК"')
    }

    foreign_banks = {
        _normalize_name('АТ "Райффайзен Банк"'),
        _normalize_name('АТ АКБ "Львів" '),
        _normalize_name('АТ "СКАЙ БАНК" '),
        _normalize_name('АТ "БТА БАНК" '),
        _normalize_name('АТ "УКРСИББАНК" '),
        _normalize_name('АТ "Ідея Банк" '),
        _normalize_name('АТ "ПРАВЕКС БАНК" '),
        _normalize_name('АТ "КРЕДІ АГРІКОЛЬ БАНК" '),
        _normalize_name('АТ "ЮНЕКС БАНК" '),
        _normalize_name('АТ "ПІРЕУС БАНК МКБ" '),
        _normalize_name('АТ "ІНГ Банк Україна" '),
        _normalize_name('АТ "ОТП БАНК" '),
        _normalize_name('АТ "СІТІБАНК" '),
        _normalize_name('АТ "ПРОКРЕДИТ БАНК" '),
        _normalize_name('АТ "УБРР"'),
        _normalize_name('АТ "НЕКСЕНТ БАНК" '),
        _normalize_name('АТ "КРЕДИТВЕСТ БАНК"'),
        _normalize_name('АТ "АГРОПРОСПЕРІС БАНК" '),
        _normalize_name('АТ "Дойче Банк ДБУ"  '),
        _normalize_name('АТ"СЕБ КОРПОРАТИВНИЙ БАНК"'),
        _normalize_name('АТ "КРЕДОБАНК"')
    }

    bank_row_by_name = {}
    for row_idx in range(6, data_last_row + 1):
        bank_name = _normalize_name(sheet.cell(row=row_idx, column=3).value)
        if bank_name:
            bank_row_by_name[bank_name] = row_idx

    state_rows = [row for name, row in bank_row_by_name.items() if name in state_banks]
    foreign_rows = [row for name, row in bank_row_by_name.items() if name in foreign_banks]
    all_rows = list(bank_row_by_name.values())

    state_row = total_row + 1
    foreign_row = total_row + 2
    private_row = total_row + 3

    category_rows = [
        (state_row, "Державні банки", state_rows),
        (foreign_row, "Банки з іноземним капіталом", foreign_rows),
        (private_row, "Банки з українським приватним капіталом", None),
    ]

    def _category_formula(rows_list, col_index):
        if not rows_list:
            return "=0"
        col_letter = get_column_letter(col_index)
        refs = ",".join([f"{col_letter}{r}" for r in sorted(rows_list)])
        return f"=IFERROR(SUM({refs}),0)"

    for out_row, category_name, source_rows in category_rows:
        sheet.cell(row=out_row, column=1).value = category_name
        sheet.cell(row=out_row, column=1).font = Font(name='Arial', size=10)
        sheet.cell(row=out_row, column=1).alignment = Alignment(horizontal='left', vertical='center')

        # Sum across only columns that have non-empty header in row 5
        for col_idx in non_empty_header_cols:
            out_cell = sheet.cell(row=out_row, column=col_idx)
            if ds_pd_col_idx is not None and col_idx == ds_pd_col_idx:
                if ds_col_idx is not None and pd_col_idx is not None:
                    ds_col_letter = get_column_letter(ds_col_idx)
                    pd_col_letter = get_column_letter(pd_col_idx)
                    out_cell.value = f"=IFERROR({ds_col_letter}{out_row}/{pd_col_letter}{out_row}*100,0)"
                else:
                    out_cell.value = "=0"
            elif ds_share_col_idx is not None and col_idx == ds_share_col_idx:
                if ds_col_idx is not None:
                    ds_col_letter = get_column_letter(ds_col_idx)
                    out_cell.value = f"=IFERROR({ds_col_letter}{out_row}/{ds_col_letter}{total_row}*100,0)"
                else:
                    out_cell.value = "=0"
            elif ovdp_share_col_idx is not None and col_idx == ovdp_share_col_idx:
                if ovdp_col_idx is not None:
                    ovdp_col_letter = get_column_letter(ovdp_col_idx)
                    out_cell.value = f"=IFERROR({ovdp_col_letter}{out_row}/{ovdp_col_letter}{total_row}*100,0)"
                else:
                    out_cell.value = "=0"
            elif source_rows is None:
                col_letter = get_column_letter(col_idx)
                out_cell.value = f"=IFERROR({col_letter}{total_row}-{col_letter}{total_row + 1}-{col_letter}{total_row + 2},0)"
            else:
                out_cell.value = _category_formula(source_rows, col_idx)
            out_cell.number_format = '#,##0.00'
            out_cell.alignment = Alignment(horizontal='center', vertical='center')

        for col_idx in range(1, last_formula_col + 1):
            cell = sheet.cell(row=out_row, column=col_idx)
            cell.border = thin_border
            cell.fill = total_fill

        sheet.row_dimensions[out_row].height = 22

    # Additional two formula tables on Index (for middle and right charts)

    # Place both small tables under the main table, starting from column B
    middle_col_name = 2  # B
    middle_col_value = 3  # C
    right_col_name = 8  # H
    right_col_value = 9  # I
    table_header_row = private_row + 2

    categories = [
        ("Державні банки", state_row),
        ("Банки з іноземним капіталом", foreign_row),
        ("Банки з українським приватним капіталом", private_row),
    ]

    # Middle table: "Частка процентів по депсертифікатах у процентних доходах банків, %"
    sheet.cell(row=table_header_row, column=middle_col_name).value = "Категорія"
    sheet.cell(row=table_header_row, column=middle_col_value).value = "ДС/ПД, %"

    # Right table: "Проценти за депозитними сертифікатами ... у розрізі банків, %"
    sheet.cell(row=table_header_row, column=right_col_name).value = "Категорія"
    sheet.cell(row=table_header_row, column=right_col_value).value = "ДС до підсумку, %"

    for idx, (category_label, category_row) in enumerate(categories, start=1):
        out_row = table_header_row + idx

        # Middle table values
        sheet.cell(row=out_row, column=middle_col_name).value = f"=A{category_row}"
        if ds_pd_col_idx is not None:
            ds_pd_letter = get_column_letter(ds_pd_col_idx)
            sheet.cell(row=out_row, column=middle_col_value).value = f"=IFERROR({ds_pd_letter}{category_row},0)"
        else:
            sheet.cell(row=out_row, column=middle_col_value).value = "=0"

        # Right table values
        sheet.cell(row=out_row, column=right_col_name).value = f"=A{category_row}"
        if ds_share_col_idx is not None:
            ds_share_letter = get_column_letter(ds_share_col_idx)
            sheet.cell(row=out_row, column=right_col_value).value = f"=IFERROR({ds_share_letter}{category_row},0)"
        else:
            sheet.cell(row=out_row, column=right_col_value).value = "=0"

    # Style both tables
    table_cols = [middle_col_name, middle_col_value, right_col_name, right_col_value]
    for r in range(table_header_row, table_header_row + 4):
        for c in table_cols:
            cell = sheet.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if r == table_header_row:
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.fill = total_fill
            else:
                if c in [middle_col_value, right_col_value]:
                    cell.number_format = '0.0'

    sheet.column_dimensions[get_column_letter(middle_col_name)].width = 38
    sheet.column_dimensions[get_column_letter(middle_col_value)].width = 14
    sheet.column_dimensions[get_column_letter(right_col_name)].width = 38
    sheet.column_dimensions[get_column_letter(right_col_value)].width = 16

    # Charts under the two small tables (pie on the left, bar on the right)
    middle_cats = Reference(sheet, min_col=middle_col_name, min_row=table_header_row + 1, max_row=table_header_row + 3)
    middle_vals = Reference(sheet, min_col=middle_col_value, min_row=table_header_row + 1, max_row=table_header_row + 3)

    bar = BarChart()
    bar.type = "col"
    bar.title = "Частка процентів по депсертифікатах у\nпроцентних доходах банків, %"
    bar.add_data(middle_vals, titles_from_data=False)
    bar.set_categories(middle_cats)
    bar.height = 13
    bar.width = 13
    bar.legend = None
    bar.x_axis.delete = False
    bar.y_axis.delete = False
    bar.x_axis.title = None
    bar.y_axis.title = None
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showSerName = False
    bar.dataLabels.showCatName = False
    bar.dataLabels.showVal = True
    bar.dataLabels.txPr = _datalabel_text_style(size_pt=13, bold=True, color="FF0000")

    right_cats = Reference(sheet, min_col=right_col_name, min_row=table_header_row + 1, max_row=table_header_row + 3)
    right_vals = Reference(sheet, min_col=right_col_value, min_row=table_header_row + 1, max_row=table_header_row + 3)

    pie_right = PieChart()
    pie_right.title = "Проценти за депозитними сертифікатами\nу розрізі банків, %"
    pie_right.title.txPr = _datalabel_text_style(size_pt=13, bold=True, color="FF0000")
    pie_right.add_data(right_vals, titles_from_data=False)
    pie_right.set_categories(right_cats)
    pie_right.height = 14
    pie_right.width = 14
    pie_right.legend = None
    pie_right.dataLabels = DataLabelList()
    pie_right.dataLabels.showSerName = False
    pie_right.dataLabels.showCatName = True
    pie_right.dataLabels.showVal = False
    pie_right.dataLabels.showPercent = True
    pie_right.dataLabels.showLeaderLines = True
    pie_right.dataLabels.txPr = _datalabel_text_style(size_pt=11, bold=True, color="000000")

    chart_top_row = table_header_row + 5
    sheet.add_chart(bar, f"B{chart_top_row}")
    sheet.add_chart(pie_right, f"H{chart_top_row}")

    print(f"  ✓ Category rows written: {len(category_rows)}")
    print("  ✓ Added 2 extra source tables on Index")

    print(f"  ✓ Totals written to row {total_row}")

    return results


def create_indicator_summary_sheet(workbook, workbook_values, index_sheet, vidpovidnist_pokaznykiv, data_last_row):
    """
    Create a formatted summary table similar to the provided example:
    - Column A: category names
    - Column B: sums in billions (SUM from Index / 1_000_000)
    - Column C: share in total, %

    "Інші процентні доходи" = "ВСЬОГО" - (ДС + ОВДП + бізнес + населення)
    """
    print("\n=== STEP 6: Creating indicator summary table ===")

    summary_sheet_name = "Indicator_Summary"

    if summary_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[summary_sheet_name])

    index_position = workbook.sheetnames.index("Index") if "Index" in workbook.sheetnames else 0
    summary_sheet = workbook.create_sheet(summary_sheet_name, index=index_position + 1)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    title_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    body_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Row 1: title
    summary_sheet.merge_cells("A1:C1")
    summary_sheet["A1"] = "ПРОЦЕНТНІ ДОХОДИ банків"
    summary_sheet["A1"].font = Font(name='Arial', size=11, bold=True)
    summary_sheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
    for col in ["A", "B", "C"]:
        summary_sheet[f"{col}1"].fill = title_fill
        summary_sheet[f"{col}1"].border = thin_border

    # Row 2: headers
    summary_sheet["B2"] = "млрд"
    summary_sheet["C2"] = "%"
    for cell_ref in ["A2", "B2", "C2"]:
        summary_sheet[cell_ref].fill = title_fill
        summary_sheet[cell_ref].font = Font(name='Arial', size=10, bold=True)
        summary_sheet[cell_ref].alignment = Alignment(horizontal='center', vertical='center')
        summary_sheet[cell_ref].border = thin_border

    # Map header text from Index row 5 to its column index
    header_to_col_idx = {}
    for col_idx in range(1, index_sheet.max_column + 1):
        header_val = index_sheet.cell(row=5, column=col_idx).value
        if header_val is not None:
            header_to_col_idx[str(header_val)] = col_idx

    last_data_row = max(6, data_last_row)

    def _to_float(value):
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.replace(' ', '').replace(',', '.')
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        return 0.0

    def _resolve_index_cell_value(cell_value):
        """Resolve direct numbers and links like ='Sheet'!$C$10 to float."""
        if isinstance(cell_value, str):
            match = re.match(r"^='(.+)'!\$?([A-Z]+)\$?(\d+)$", cell_value)
            if match:
                src_sheet_name = match.group(1)
                src_col = match.group(2)
                src_row = int(match.group(3))
                print(f"ids: {src_col}, {src_row}")

                print(src_sheet_name, src_sheet_name in workbook.sheetnames)
                if src_sheet_name in workbook.sheetnames:
                    src_sheet = workbook[src_sheet_name]
                    src_raw = src_sheet[f"{src_col}{src_row}"].value
                    print(f"raw: {src_raw}")

                    # Prefer cached calculated value (data_only workbook)
                    print(src_sheet_name in workbook_values.sheetnames)
                    if src_sheet_name in workbook_values.sheetnames:
                        src_values_sheet = workbook_values[src_sheet_name]
                        print(src_values_sheet)
                        src_cached = src_values_sheet[f"{src_col}{src_row}"].value
                        print(f"cached: {src_cached}")
                        if src_cached is not None:
                            # print(f"to_float: {_to_float(src_cached)}")
                            return _to_float(src_cached)
                        print(f"to_float: {_to_float(src_cached)}")
                    return _to_float(src_raw)
                return 0.0

        return _to_float(cell_value)

    def indicator_sum_value(header_name):
        col_idx = header_to_col_idx.get(header_name)
        print(header_name, col_idx)
        if col_idx is None:
            return 0.0

        # total = 0.0
        # for row_idx in range(6, last_data_row + 1):
        #     raw_value = index_sheet.cell(row=row_idx, column=col_idx).value
        #     val = _resolve_index_cell_value(raw_value)
        #     total += val
        #     print(raw_value, val, total)
        #
        # return total / 1_000_000
        col = get_column_letter(col_idx)
        # start_cell = f"{index_sheet.title}!${col}$6"
        # end_cell = f"{index_sheet.title}!${col}${last_data_row}"
        total = f"=SUM({index_sheet.title}!{col}6:{col}{last_data_row})/1000000"

        return total

    ds_header = "ПРОЦЕНТИ ЗА ДЕПОЗИТНИМИ СЕРТИФІКАТАМИ (ДС), тис. грн"
    ovdp_header = "ПРОЦЕНТИ ЗА ОВДП (ОВДП), тис.грн"
    business_header = "ПРОЦЕНТИ ЗА КРЕДИТАМИ БІЗНЕСУ"
    population_header = "ПРОЦЕНТИ ЗА КРЕДИТАМИ НАСЕЛЕННЮ"
    total_header = "ПРОЦЕНТНІ ДОХОДИ ВСЬОГО (ПД), тис. грн"

    ds_value = indicator_sum_value(ds_header)
    ovdp_value = indicator_sum_value(ovdp_header)
    business_value = indicator_sum_value(business_header)
    population_value = indicator_sum_value(population_header)
    total_value = indicator_sum_value(total_header)
    # other_value = total_value - (ds_value + ovdp_value + business_value + population_value)
    other_value = (f"={summary_sheet.cell(row=8, column=2).coordinate} - ("
                   f"{summary_sheet.cell(row=3, column=2).coordinate} +"
                   f"{summary_sheet.cell(row=4, column=2).coordinate} +"
                   f"{summary_sheet.cell(row=5, column=2).coordinate} +"
                   f"{summary_sheet.cell(row=6, column=2).coordinate})")

    rows = [
        (3, "Депозитні сертифікати", ds_value),
        (4, "ОВДП", ovdp_value),
        (5, "Кредити бізнесу", business_value),
        (6, "Кредити населенню", population_value),
        (7, "Інші процентні доходи", other_value),
        (8, "ВСЬОГО", total_value),
    ]

    for row_idx, row_name, numeric_value in rows:
        summary_sheet.cell(row=row_idx, column=1).value = row_name
        summary_sheet.cell(row=row_idx, column=2).value = numeric_value

        if row_idx == 8:
            summary_sheet.cell(row=row_idx, column=3).value = 100.0
        else:
        #     if total_value == 0:
        #         summary_sheet.cell(row=row_idx, column=3).value = 0.0
        #     else:
        #     summary_sheet.cell(row=row_idx, column=3).value = float(numeric_value) / total_value * 100.0
            num = summary_sheet.cell(row=row_idx, column=2).coordinate
            tot = summary_sheet.cell(row=8, column=2).coordinate
            summary_sheet.cell(row=row_idx, column=3).value = f"={num}/{tot} * 100"

        for col_idx in range(1, 4):
            cell = summary_sheet.cell(row=row_idx, column=col_idx)
            cell.fill = body_fill
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.0'

    # Optional last row from sample layout
    summary_sheet.cell(row=9, column=1).value = "у % до доходів банків"
    summary_sheet.cell(row=9, column=2).value = ""
    summary_sheet.cell(row=9, column=3).value = ""
    for col_idx in range(1, 4):
        cell = summary_sheet.cell(row=9, column=col_idx)
        cell.fill = body_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')

    # Emphasize total row
    for col_idx in range(1, 4):
        summary_sheet.cell(row=8, column=col_idx).font = Font(name='Arial', size=10, bold=True)

    summary_sheet.column_dimensions['A'].width = 38
    summary_sheet.column_dimensions['B'].width = 12
    summary_sheet.column_dimensions['C'].width = 10
    summary_sheet.column_dimensions['E'].width = 4
    summary_sheet.column_dimensions['F'].width = 20
    summary_sheet.column_dimensions['G'].width = 20
    summary_sheet.column_dimensions['H'].width = 20
    summary_sheet.row_dimensions[1].height = 22

    # Pie chart based on rows 3-7 from Indicator_Summary table
    labels = Reference(summary_sheet, min_col=1, min_row=3, max_row=7)
    data = Reference(summary_sheet, min_col=2, min_row=2, max_row=7)

    pie = PieChart()
    pie.title = "Структура процентних доходів банків\nза 12 міс. 2025 року, %"
    pie.title.txPr = _datalabel_text_style(size_pt=13, bold=True, color="FF0000")
    pie.title.overlay = False
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 10
    pie.width = 13
    pie.legend = None

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showSerName = False
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showPercent = True
    pie.dataLabels.showLeaderLines = True
    pie.dataLabels.txPr = _datalabel_text_style(size_pt=10, bold=True, color="000000")

    summary_sheet.add_chart(pie, "F1")

    print(f"  ✓ Created '{summary_sheet_name}' in sample layout")
    return summary_sheet


def process_excel(name_of_the_file):
    """Main execution function."""
    print("=" * 60)
    print("COMBINED BANK DATA PROCESSING SCRIPT")
    print("=" * 60)

    # Configuration
    # name_of_the_file = "OSB_bank_2026-01-01 (5).xlsx"
    # name_of_the_file = "TEST.xlsx"

    # Mapping of column headers to indicator names
    # Valid parameter values: 1-9 (debit_total, debit_nc, debit_ic, credit_total, credit_nc, credit_ic, balance_total, balance_nc, balance_ic)


    vidpovidnist_pokaznykiv = {
        "ПРОЦЕНТИ ЗА ДЕПОЗИТНИМИ СЕРТИФІКАТАМИ (ДС), тис. грн": ("Проценти по ДС (коди 6127-6128)", 7),
        "ПРОЦЕНТНІ ДОХОДИ ВСЬОГО (ПД), тис. грн": ("Процентні доходи (коди р.60,р.61)", 7),
        "ПРОЦЕНТИ ЗА ОВДП (ОВДП), тис.грн": ("Проценти по ОВДП (коди 6120-6122)", 7),
        "ПРОЦЕНТИ ЗА КРЕДИТАМИ БІЗНЕСУ": ("Проценти за кредитами НК (коди 602, 603, 609)", 7),
        "ПРОЦЕНТИ ЗА КРЕДИТАМИ НАСЕЛЕННЮ": ("Проценти за кредитами ДГ (коди 605, 606, 611)", 7),
        "ПРОЦЕНТИ  ЗА КРЕДИТАМИ ЗДУ": ("Проценти за кредитами ЗДУ (код 604)", 7),
        "ПРОЦЕНТНІ ВИТРАТИ БАНКІВ ЗА ДЕПОЗИТАМИ ТА НАСЕЛЕННЯ, тис.грн": ("Процентні витрати банків ДГ - всього ", 7),
        "ДС/ПД, %": ("Проценти по ДС у % до процентних доходів", 7),
        "(ДС+ОВДП)/ПД, %": ("Проценти по ДС+ОВДП у % до процентних доходів", 7),
        "ОВДП/ПД, %": ("Проценти по ОВДП у % до процентних доходів", 7),
        "Обсяг ДС дебит (середні, поділено на кількість банк.днів)": (
            "Обсяг ДС дебит (середні, поділено на кількість банк.днів) (код 1430, 1440)", 1),
        "Обсяг ДС кредит (середні, поділено на кількість банк.днів)": (
            "Обсяг ДС кредит (середні, поділено на кількість банк.днів) (код 1430, 1440)", 5),
        "Кошти в НБУ (кор.рахунок) (кредит) (код 1200)":("Кошти в НБУ (кор.рахунок) (кредит) (код 1200)",4),
        "Кошти в НБУ (УСЬОГО) (кредит) (розділ 12)":("Кошти в НБУ (УСЬОГО) (кредит) (розділ 12)",4),
        "Проценти за поточними вкладами ДГ (код 7040)":("Проценти за поточними вкладами ДГ (код 7040)",7),
        "Проценти за строковими вкладами ДГ (код 7041)":("Проценти за строковими вкладами ДГ (код 7041)",7),
        "Кредити субʼєктам господарювання (р.20, р.23, 260)":("Кредити субʼєктам господарювання (р.20, р.23, 260)",7),
        "Кредити ЗДУ (р.21)":("Кредити ЗДУ (р.21)",7),
        "Кредити фізичним особам (р. 22, 24, 262)":("Кредити фізичним особам (р. 22, 24, 262)",7),
        "Кредити небанківським установам (265)":("Кредити небанківським установам (265)",7),
        "Прибуток звітного року (5040)":("Прибуток звітного року (5040)",7),
        "Збиток звітного року (5041)":("Збиток звітного року (5041)",7),
        "Податок на прибуток (7900)":("Податок на прибуток (7900)",7),
        "Інші податки (741)":("Інші податки (741)",7),
        "Відрахування в резерви (770)":("Відрахування в резерви (770)",7),
        "Загальні доходи (6)":("Загальні доходи (6)",7),
        "Загальні витрати (7)":("Загальні витрати (7)",7)
    }

    # STEP 0: Load Excel file
    print(f"\n=== STEP 0: Loading Excel file ===")
    print(f"  File: {name_of_the_file}")
    wb = xl.load_workbook(name_of_the_file)
    wb_values = xl.load_workbook(name_of_the_file, data_only=True)
    print(f"  ✓ Loaded successfully! Total sheets: {len(wb.sheetnames)}")

    # STEP 1: Create Index sheet and format it
    print(f"\n=== STEP 1: Creating and formatting Index sheet ===")

    # Find matching sheets
    matching_sheets = []
    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue
        if is_valid_sheet_name(sheet_name):
            matching_sheets.append(sheet_name)

    # Sort by bank code
    matching_sheets.sort(key=lambda x: (int(re.match(r'^\s*(\d+)', x).group(1)), x))
    print(f"  ✓ Found {len(matching_sheets)} bank sheets")

    # Create and format Index sheet
    index_sheet = create_index_sheet(wb)
    format_header(index_sheet)
    add_hyperlinks(wb, index_sheet, matching_sheets)

    # STEP 2-4: Fill formulas (from Danya's code)
    fill_formulas(wb, vidpovidnist_pokaznykiv, name_of_the_file)

    # STEP 5: Calculate sums
    data_last_row = index_sheet.max_row
    column_sums = sum_columns_by_header(wb, wb_values, index_sheet, data_last_row)

    # STEP 6: Create summary table
    create_indicator_summary_sheet(wb, wb_values, index_sheet, vidpovidnist_pokaznykiv, data_last_row)

    # STEP 7: Save file
    print(f"\n=== STEP 7: Saving file ===")
    print(f"  Saving to: {name_of_the_file}")
    wb.save(name_of_the_file)
    print(f"  ✓ File saved successfully!")

    print("\n" + "=" * 60)
    print("✓✓✓ PROCESSING COMPLETE! ✓✓✓")
    print("=" * 60)
    print(f"Summary:")
    print(f"  - Created Index sheet")
    print(f"  - Added {len(matching_sheets)} bank sheet links")
    print(f"  - Filled formulas for indicators")
    print(f"  - Calculated {len(column_sums)} column sums")
    print(f"  - Created indicator summary table")
    print("=" * 60)


if __name__ == "__main__":
    process_excel()
