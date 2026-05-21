from .utils import *
from .utils import _datalabel_text_style, _normalize_name
from .indicator_tables_creator import IndicatorTablesCreator
from ..account_rows import BalanceSheet
import time
import re
import json
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList


class IndexSheetCreator:
    # UPDATED MAPPING: "Index Header" -> ("Indicator Name from CSV", "Alias Column in Target Table")
    INDICATOR_CORRESPONDENCE = {
        "ПРОЦЕНТИ ЗА ДЕПОЗИТНИМИ СЕРТИФІКАТАМИ (ДС), тис. грн": ("Проценти по ДС (коди 6127-6128)", "Balance_Total"),
        "ПРОЦЕНТНІ ДОХОДИ ВСЬОГО (ПД), тис. грн": ("Процентні доходи (коди р.60,р.61)", "Balance_Total"),
        "ПРОЦЕНТИ ЗА ОВДП (ОВДП), тис.грн": ("Проценти по ОВДП (коди 6120-6122)", "Balance_Total"),
        "ПРОЦЕНТИ ЗА КРЕДИТАМИ БІЗНЕСУ": ("Проценти за кредитами НК (коди 602, 603, 609)", "Balance_Total"),
        "ПРОЦЕНТИ ЗА КРЕДИТАМИ НАСЕЛЕННЮ": ("Проценти за кредитами ДГ (коди 605, 606, 611)", "Balance_Total"),
        "ПРОЦЕНТИ  ЗА КРЕДИТАМИ ЗДУ": ("Проценти за кредитами ЗДУ (код 604)", "Balance_Total"),
        "ПРОЦЕНТНІ ВИТРАТИ БАНКІВ ЗА ДЕПОЗИТАМИ ТА НАСЕЛЕННЯ, тис.грн": (
            "Процентні витрати банків ДГ - всього ", "Balance_Total"),
        "ДС/ПД, %": ("Проценти по ДС у % до процентних доходів", "Balance_Total"),
        "(ДС+ОВДП)/ПД, %": ("Проценти по ДС+ОВДП у % до процентних доходів", "Balance_Total"),
        "ОВДП/ПД, %": ("Проценти по ОВДП у % до процентних доходів", "Balance_Total"),
        "Обсяг ДС": (
            "Обсяг ДС (код 1430, 1440)", "Balance_Total"),
        "Кошти в НБУ (кор.рахунок) (кредит) (код 1200)": (
            "Кошти в НБУ (кор.рахунок) (кредит) (код 1200)", "Credit_Total"),
        "Кошти в НБУ (УСЬОГО) (кредит) (розділ 12)": ("Кошти в НБУ (УСЬОГО) (кредит) (розділ 12)", "Credit_Total"),
        "Проценти за поточними вкладами ДГ (код 7040)": (
        "Проценти за поточними вкладами ДГ (код 7040)", "Balance_Total"),
        "Проценти за строковими вкладами ДГ (код 7041)": (
            "Проценти за строковими вкладами ДГ (код 7041)", "Balance_Total"),
        "Кредити субʼєктам господарювання (р.20, р.23, 260)": (
            "Кредити субʼєктам господарювання (р.20, р.23, 260)", "Balance_Total"),
        "Кредити ЗДУ (р.21)": ("Кредити ЗДУ (р.21)", "Balance_Total"),
        "Кредити фізичним особам (р. 22, 24, 262)": ("Кредити фізичним особам (р. 22, 24, 262)", "Balance_Total"),
        "Кредити небанківським установам (265)": ("Кредити небанківським установам (265)", "Balance_Total"),
        "Прибуток звітного року (5040)": ("Прибуток звітного року (5040)", "Balance_Total"),
        "Збиток звітного року (5041)": ("Збиток звітного року (5041)", "Balance_Total"),
        "Податок на прибуток (7900)": ("Податок на прибуток (7900)", "Balance_Total"),
        "Інші податки (741)": ("Інші податки (741)", "Balance_Total"),
        "Відрахування в резерви (770)": ("Відрахування в резерви (770)", "Balance_Total"),
        "Загальні доходи (6)": ("Загальні доходи (6)", "Balance_Total"),
        "Загальні витрати (7)": ("Загальні витрати (7)", "Balance_Total")
    }

    def __init__(self, indicator_tables_creator):
        print(f"\n=== STAGE 3: Index Sheet Creation ===")
        print(f"    Initialising an IndexSheetCreator object")
        self.indicator_tables_creator = indicator_tables_creator
        self.workbook = self.indicator_tables_creator.workbook
        self.workbook_data = getattr(self.indicator_tables_creator, 'workbook_data', self.workbook)
        self.parent_file = getattr(self.indicator_tables_creator, 'parent_file', None)
        self.banks_file = self.indicator_tables_creator.banks_file
        self.progress_var = getattr(self.indicator_tables_creator, 'progress_var', None)
        self.delay = 0.01
        self.percentage = 5

        self.index_sheet = self.create_index_sheet()
        self.matching_sheets = self.find_matching_sheets()
        self.sheets_dict = self.indicator_tables_creator.sheets_dict

        self.config()
        print(f"    STAGE 3 COMPLETE")

    def update_progress(self, value):
        if self.progress_var:
            current = self.progress_var.get()
            new = current + value / 100.0
            self.progress_var.set(new)
            time.sleep(self.delay)

    # def _get_bank_category(self, bank_name):
    #     """Helper that loads the JSON to categorize the bank name for the new column"""
    #
    #     # AGGRESSIVE CLEANER: Removes ALL spaces, quotes (all types), punctuation, etc.
    #     def clean_for_match(name):
    #         if not name: return ""
    #         # Strip anything that isn't a letter or number
    #         s = re.sub(r'[\s\-\,\.\(\)\'\"«»“”`]', '', str(name).lower())
    #         # Normalize common Cyrillic/Latin lookalikes just in case
    #         return s.translate(str.maketrans('ііїєсхрАТМВНЕО', 'iiiеcxpATMBHEO'))
    #
    #     norm_name = clean_for_match(bank_name)
    #
    #     if not hasattr(self, '_state_banks_set'):
    #         self._state_banks_set = set()
    #         self._foreign_banks_set = set()
    #         self._private_banks_set = set()
    #         # try:
    #         #     if self.banks_file:
    #         #         with open(self.banks_file, 'r', encoding='utf-8') as f:
    #         #             cat_data = json.load(f)
    #         #
    #         #         # Store aggressively cleaned versions of the JSON names
    #         #         for b in cat_data.get("groups", {}).get("state_owned_banks", {}).get("banks", []):
    #         #             self._state_banks_set.add(clean_for_match(b.get("name", "")))
    #         #         for b in cat_data.get("groups", {}).get("foreign_banking_groups", {}).get("banks", []):
    #         #             self._foreign_banks_set.add(clean_for_match(b.get("name", "")))
    #         #         for b in cat_data.get("groups", {}).get("private_capital_banks", {}).get("banks", []):
    #         #             self._private_banks_set.add(clean_for_match(b.get("name", "")))
    #         # except Exception as e:
    #         #     pass
    #         try:
    #             if self.banks_file:
    #                 cat_data = pd.read_excel(self.banks_file, index_col=0)
    #
    #                 self._state_banks_set = set(cat_data[cat_data["CLASS"] == "державний"]["NAME"])
    #                 self._foreign_banks_set = set(cat_data[cat_data["CLASS"] == "іноземний"]["NAME"])
    #                 self._private_banks_set = set(cat_data[cat_data["CLASS"] == "приватний"]["NAME"])
    #                 # Store aggressively cleaned versions of the JSON names
    #                 # for b in cat_data[cat_data["CLASS"] == "державний"]["NAME"]:
    #                 #     self._state_banks_set.add(clean_for_match(b.("name", "")))
    #                 # for b in cat_data.get("groups", {}).get("foreign_banking_groups", {}).get("banks", []):
    #                 #     self._foreign_banks_set.add(clean_for_match(b.get("name", "")))
    #                 # for b in cat_data.get("groups", {}).get("private_capital_banks", {}).get("banks", []):
    #                 #     self._private_banks_set.add(clean_for_match(b.get("name", "")))
    #         except Exception as e:
    #             pass
    #
    #     # Check if the cleaned sheet name is a substring of the cleaned JSON name (or vice versa)
    #     if norm_name in self._state_banks_set or any(
    #             s in norm_name or norm_name in s for s in self._state_banks_set if s):
    #         return "Державні банки"
    #     if norm_name in self._foreign_banks_set or any(
    #             f in norm_name or norm_name in f for f in self._foreign_banks_set if f):
    #         return "Банки з іноземним капіталом"
    #     if norm_name in self._private_banks_set or any(
    #             f in norm_name or norm_name in f for f in self._private_banks_set if f):
    #         return "Банки з українським приватним капіталом"
    #
    #     return "—"

    def _get_bank_category(self, bank_name):
        """Helper that loads the XLSX to categorize the bank using its code or name"""
        import re
        import pandas as pd

        # 1. Extract numerical bank code if available in the title/string (e.g., "46 ПРИВАТБАНК" -> 46)
        bank_code = ""
        match = re.match(r'^\s*(\d+)', str(bank_name))
        if match:
            bank_code = match.group(1)
        print(f"code: {bank_code}, name: {bank_name}")

        # 2. Aggressive string cleaner for fallback matching
        def clean_for_match(name):
            if not name: return ""
            s = re.sub(r'[\s\-\,\.\(\)\'\"«»“”`—]', '', str(name).lower())
            return s.translate(str.maketrans('ііїєсхрАТМВНЕО', 'iiiеcxpATMBHEO'))

        norm_name = clean_for_match(bank_name)

        # 3. Lazy initialization of the database cache maps
        if not hasattr(self, '_bank_code_map'):
            self._bank_code_map = {}  # {nkb_int: class_str}
            self._state_banks_set = set()
            self._foreign_banks_set = set()
            self._private_banks_set = set()

            try:
                if self.banks_file:
                    # Load without forcing index columns blindly
                    cat_data = pd.read_excel(self.banks_file)

                    # print(cat_data.head(25))
                    # Normalize column headers to lowercase & strip whitespace
                    cat_data.columns = cat_data.columns.str.lower().str.strip()
                    # print(cat_data.columns)

                    if 'class' in cat_data.columns:
                        # Build Bulletproof Code Map if NKB column exists
                        if 'nkb' in cat_data.columns:
                            cat_data['nkb_numeric'] = pd.to_numeric(cat_data['nkb'], errors='coerce')
                            for _, row in cat_data.dropna(subset=['nkb_numeric']).iterrows():
                                self._bank_code_map[int(row['nkb_numeric'])] = str(row['class']).strip().lower()

                        # Build Cleaned Text Sets for Fallback Lookup
                        if 'name' in cat_data.columns:
                            for _, row in cat_data.iterrows():
                                cls = str(row['class']).strip().lower()
                                cleaned_name = clean_for_match(row['name'])
                                if not cleaned_name: continue

                                if cls == "державний":
                                    self._state_banks_set.add(cleaned_name)
                                elif cls == "іноземний":
                                    self._foreign_banks_set.add(cleaned_name)
                                elif cls == "приватний":
                                    self._private_banks_set.add(cleaned_name)
            except Exception as e:
                print(f"Warning: Failed to cache bank classification Excel: {e}")

        # ===================================================
        # LOOKUP LAYER 1: Match precisely by numeric NKB code
        # ===================================================
        if bank_code:
            try:
                target_nkb = int(bank_code)
                if target_nkb in self._bank_code_map:
                    mapped_class = self._bank_code_map[target_nkb]
                    if mapped_class == "державний": return "Державний"
                    if mapped_class == "іноземний": return "Іноземний капітал"
                    if mapped_class == "приватний": return "Приватний капітал"
            except ValueError:
                pass

        # ===================================================
        # LOOKUP LAYER 2: Fallback to aggressive text matching
        # ===================================================
        if norm_name in self._state_banks_set or any(
                s in norm_name or norm_name in s for s in self._state_banks_set if s):
            return "Державний"

        if norm_name in self._foreign_banks_set or any(
                f in norm_name or norm_name in f for f in self._foreign_banks_set if f):
            return "Іноземний капітал"

        if norm_name in self._private_banks_set or any(
                p in norm_name or norm_name in p for p in self._private_banks_set if p):
            return "Приватний капітал"

        return "Інше"

    def config(self):
        increment_percent = self.percentage / 4

        print(f"    Performing Index Sheet configuration:")
        self.format_header();
        self.update_progress(increment_percent)
        self.add_hyperlinks();
        self.update_progress(increment_percent)
        self.fill_formulas();
        self.update_progress(increment_percent)
        self.data_last_row = self.index_sheet.max_row
        self.sum_columns_by_header();
        self.update_progress(increment_percent)
        print(f"    Index Sheet configured")

    def find_matching_sheets(self):
        # RESTORED: Your exact original regex to properly collect the bank sheets
        matching_sheets = []
        for sheet_name in self.workbook.sheetnames:
            if sheet_name == "Index":
                continue
            if re.match(r'^\s*(\d+)', sheet_name):
                matching_sheets.append(sheet_name)
        matching_sheets.sort(key=lambda x: (int(re.match(r'^\s*(\d+)', x).group(1)), x))
        print(f"    Valid sheet names separated: {len(matching_sheets)} found.")
        return matching_sheets

    def create_index_sheet(self):
        if "Index" in self.workbook.sheetnames:
            print(f"    'Index' sheet already exists - removing it...")
            index_sheet = self.workbook["Index"]
            self.workbook.remove(index_sheet)
            print(f"     ✓ Existing Index sheet removed")

        print(f"    Creating new Index sheet...")
        index_sheet = self.workbook.create_sheet("Index", 0)
        print(f"    ✓ Index sheet created")
        return index_sheet

    def format_header(self):
        """Format the header rows of the Index sheet. (UNTOUCHED)"""
        print(f"        Formatting Index Sheet headers...")
        for row_id in [1, 2, 3, 4]:
            self.index_sheet.row_dimensions[row_id].hidden = True

        self.index_sheet['C2'] = "Оборотно-сальдовий баланс банку*"
        self.index_sheet['D2'] = "відповідно до Додатку 1 до постанови Правління Національного банку України"
        self.index_sheet[
            'D3'] = "від 15 березня 2018 року № 11 \"Про встановлення переліку інформації, що підлягає обов'язковому опублікуванню банками України\" (зі змінами)"
        self.index_sheet[
            'D4'] = "за даними статистичної звітності з файлів 02X \"Дані про обороти та залишки на рахунках\"."

        self.index_sheet['A5'] = "NKB"
        self.index_sheet['B5'] = "Порядковий номер"
        self.index_sheet['C5'] = "Назва банку"
        self.index_sheet['D5'] = "Категорія банку"  # SHIFTED: NEW COLUMN
        self.index_sheet['E5'] = "МФО"  # SHIFTED: Moved from D to E

        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col in ['A5', 'B5', 'C5', 'D5', 'E5']:
            self.index_sheet[col].font = Font(name='Arial', size=11, bold=True, color="000000")
            self.index_sheet[col].alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            self.index_sheet[col].border = thin_border

        self.index_sheet.column_dimensions['A'].width = 15
        self.index_sheet.column_dimensions['B'].width = 12
        self.index_sheet.column_dimensions['C'].width = 30
        self.index_sheet.column_dimensions['D'].width = 35  # Give enough space for category text
        self.index_sheet.column_dimensions['E'].width = 15

        headers_e_ad = {
            6: "ПРОЦЕНТИ ЗА ДЕПОЗИТНИМИ СЕРТИФІКАТАМИ (ДС), тис. грн",
            7: "ПРОЦЕНТИ ЗА ОВДП (ОВДП), тис.грн",
            8: "ПРОЦЕНТИ  ЗА КРЕДИТАМИ ЗДУ",
            9: "ПРОЦЕНТИ ЗА КРЕДИТАМИ БІЗНЕСУ",
            10: "ПРОЦЕНТИ ЗА КРЕДИТАМИ НАСЕЛЕННЮ",
            11: "ПРОЦЕНТНІ ДОХОДИ ВСЬОГО (ПД), тис. грн",
            12: "ДС/ПД, %",
            13: "ДС до підсумку, %",
            14: "(ДС+ОВДП)/ПД, %",
            15: "ОВДП/ПД, %",
            16: "ОВДП до підсумку",
            17: "Обсяг ДС",
            18: "Кошти в НБУ (кор.рахунок) (кредит) (код 1200)",
            19: "Кошти в НБУ (УСЬОГО) (кредит) (розділ 12)",
            20: "Проценти за поточними вкладами ДГ (код 7040)",
            21: "Проценти за строковими вкладами ДГ (код 7041)",
            22: "Кредити субʼєктам господарювання (р.20, р.23, 260)",
            23: "Кредити ЗДУ (р.21)",
            24: "Кредити фізичним особам (р. 22, 24, 262)",
            25: "Кредити небанківським установам (265)",
            26: "Прибуток звітного року (5040)",
            27: "Збиток звітного року (5041)",
            28: "Податок на прибуток (7900)",
            29: "Інші податки (741)",
            30: "Відрахування в резерви (770)",
            31: "Загальні доходи (6)",
            32: "Загальні витрати (7)"
        }

        # SHIFTED: Data starts from 6
        for col_idx in range(6, 12):
            col_letter = get_column_letter(col_idx)
            cell = self.index_sheet[f'{col_letter}5']
            cell.value = headers_e_ad.get(col_idx, "")
            cell.font = Font(name='Arial', size=9, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            self.index_sheet.column_dimensions[col_letter].width = 18
            self.index_sheet.row_dimensions[5].height = 60

        for col_idx in range(12, 41):
            col_letter = get_column_letter(col_idx)
            cell = self.index_sheet[f'{col_letter}5']
            cell.value = headers_e_ad.get(col_idx, "")
            cell.font = Font(name='Arial', size=9, bold=True, color="000000")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            self.index_sheet.column_dimensions[col_letter].width = 18
            self.index_sheet.row_dimensions[5].height = 60

        self.index_sheet.freeze_panes = "F6"  # SHIFTED to freeze new category column
        print(f"        ✓ Index Sheet headers formatted")

    def add_hyperlinks(self):
        print(f"        Setting up hyperlinks...")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        current_row = 6
        index_counter = 1

        for sheet_name in self.matching_sheets:
            # 1. EXTRACT NKB AND NAME
            # We use the regex match to get the NKB (e.g., '695') from the sheet name
            match = re.match(r'^\s*(\d+)', sheet_name)
            nkb_str = match.group(1) if match else ""
            text_part = sheet_name[match.end():].strip() if match else sheet_name
            print(f"sheet name: {sheet_name}, match: {match}, nkb: {nkb_str}, text part: {text_part}")

            # 2. FILL COLUMN A (NKB)
            cell_a = self.index_sheet.cell(row=current_row, column=1)
            cell_a.value = nkb_str
            cell_a.hyperlink = f"#'{sheet_name}'!A1"
            cell_a.hyperlink.target = None
            cell_a.hyperlink.location = f"'{sheet_name}'!A1"
            cell_a.font = Font(name='Arial', size=11, color="0000FF", underline="single")
            cell_a.border = thin_border
            cell_a.alignment = Alignment(horizontal='center', vertical='center')

            # 3. FILL COLUMN B (ORDINAL)
            cell_b = self.index_sheet.cell(row=current_row, column=2)
            cell_b.value = index_counter
            cell_b.font = Font(name='Arial', size=11)
            cell_b.border = thin_border
            cell_b.alignment = Alignment(horizontal='center', vertical='center')

            # 4. FILL COLUMN C (BANK NAME)
            cell_c = self.index_sheet.cell(row=current_row, column=3)
            cell_c.value = text_part
            cell_c.hyperlink = f"#'{sheet_name}'!A1"
            cell_c.hyperlink.target = None
            cell_c.hyperlink.location = f"'{sheet_name}'!A1"
            cell_c.font = Font(name='Arial', size=11, color="0000FF", underline="single")
            cell_c.border = thin_border
            cell_c.alignment = Alignment(horizontal='left', vertical='center')

            # 5. FILL COLUMN D (CATEGORY)
            category = self._get_bank_category(sheet_name.strip())
            cell_d = self.index_sheet.cell(row=current_row, column=4)
            cell_d.value = category
            cell_d.font = Font(name='Arial', size=11)
            cell_d.border = thin_border
            cell_d.alignment = Alignment(horizontal='center', vertical='center')

            # 6. LOOKUP AND FILL COLUMN E (MFO)
            mfo_value = None
            if "Зміст" in self.workbook.sheetnames and nkb_str:
                contents = self.workbook["Зміст"]
                # Standard NBU 02X format: Col 1 = NKB, Col 3 = MFO
                for row_idx in range(1, contents.max_row + 1):
                    cell_val = contents.cell(row=row_idx, column=1).value
                    if cell_val is None: continue
                    try:
                        cell_text = str(cell_val).strip()
                        # Match NKB from sheet name to NKB in contents
                        if cell_text == nkb_str or (cell_text.isdigit() and int(cell_text) == int(nkb_str)):
                            mfo_value = contents.cell(row=row_idx, column=3).value
                            break
                    except Exception:
                        pass

            cell_e = self.index_sheet.cell(row=current_row, column=5)
            if mfo_value is not None:
                cell_e.value = str(mfo_value).strip()
                cell_e.hyperlink = f"#'{sheet_name}'!A1"
                cell_e.hyperlink.target = None
                cell_e.hyperlink.location = f"'{sheet_name}'!A1"
                cell_e.font = Font(name='Arial', size=11, color="0000FF", underline="single")
            cell_e.border = thin_border
            cell_e.alignment = Alignment(horizontal='center', vertical='center')

            current_row += 1
            index_counter += 1

        for row in range(6, current_row):
            self.index_sheet.row_dimensions[row].height = 22

        print(f"        ✓ Corrected NKB/MFO mapping for {len(self.matching_sheets)} bank sheets")

    def fill_formulas(self):
        print(f"        Filling out Index Sheet formulae...")
        ws_index = self.index_sheet
        ind_cor = self.INDICATOR_CORRESPONDENCE

        # AGGRESSIVE CLEANER: Removes ALL spaces, quotes, hyphens, parentheses to make string-matching unshakeable
        def normalize_str(s):
            if not s: return ""
            return re.sub(r'[\s\-\,\.\(\)\'\"«»]', '', str(s).lower())

        sheets_dict = getattr(self.indicator_tables_creator, 'sheets_dict', {})

        # Iterate safely over the extracted matching sheets, bypassing tricky hyperlink string parsing
        for idx, sheet_name in enumerate(self.matching_sheets):
            row_idx = 6 + idx

            # 1. BULLETPROOF OBJECT RETRIEVAL: Check if the sheet title literally matches to bypass key weirdness
            bs_object = None
            for obj in sheets_dict.values():
                if getattr(obj.sheet, 'title', None) == sheet_name:
                    bs_object = obj
                    break

            if not bs_object:
                print(f"            ⚠ Missing BalanceSheet object for {sheet_name}. Formula left as 0.")
                continue

            start_coord = getattr(bs_object, 'start_cell_coord', None)
            if not start_coord:
                continue

            target_ws = self.workbook[sheet_name]
            col_match = re.match(r"([A-Z]+)(\d+)", start_coord)
            if not col_match: continue

            table_anchor_col = column_index_from_string(col_match.group(1))
            table_anchor_row = int(col_match.group(2))

            # 2. Map Columns (e.g., 'Credit_Total' -> 'AA')
            col_map = {}
            for c_idx in range(table_anchor_col, target_ws.max_column + 1):
                val = target_ws.cell(row=table_anchor_row, column=c_idx).value
                if val:
                    col_map[str(val).strip()] = get_column_letter(c_idx)

            # 3. Map Rows using the aggressive cleaner
            row_map = {}
            empty_count = 0
            for r_idx in range(table_anchor_row + 1, table_anchor_row + 200):
                ind_name = target_ws.cell(row=r_idx, column=table_anchor_col + 1).value
                if not ind_name:
                    empty_count += 1
                    if empty_count > 10: break
                    continue
                empty_count = 0
                row_map[normalize_str(ind_name)] = r_idx

            # Safeguard the sheet name in the Excel reference formula just in case of weird quotes
            safe_sheet_name = sheet_name.replace("'", "''")

            # 4. Write Formulas Row by Row - SHIFTED: Iterates from column 6
            for col_idx in range(6, ws_index.max_column + 1):
                index_header = ws_index.cell(row=5, column=col_idx).value
                if not index_header: continue

                index_header_str = str(index_header).strip()

                if index_header_str in ind_cor:
                    target_name, target_alias = ind_cor[index_header_str]
                    clean_target = normalize_str(target_name)

                    target_row_num = row_map.get(clean_target)

                    # If direct clean-match fails, try substring (in case CSV has extra info)
                    if not target_row_num:
                        for r_clean, r_idx in row_map.items():
                            if clean_target in r_clean or r_clean in clean_target:
                                target_row_num = r_idx
                                break

                    if target_row_num and target_alias in col_map:
                        target_col_letter = col_map[target_alias]
                        ws_index.cell(row=row_idx,
                                      column=col_idx).value = f"='{safe_sheet_name}'!{target_col_letter}{target_row_num}"
                    else:
                        ws_index.cell(row=row_idx, column=col_idx).value = 0

        print(f"            ✓ All formulae filled out successfully!")

    def sum_columns_by_header(self):
        """Write totals row under data using Excel formulas."""
        print(f"        Creating the totals rows")
        results = {}
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        print(f"            Calculating the totals for each column")

        total_row = self.data_last_row + 1
        self.index_sheet.cell(row=total_row, column=1).value = "Всього"
        self.index_sheet.cell(row=total_row, column=1).font = Font(name='Arial', size=10, bold=True)
        self.index_sheet.cell(row=total_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

        non_empty_header_cols = []
        header_to_col = {}

        # SHIFTED: Evaluates data starting at column 6
        for col_idx in range(6, self.index_sheet.max_column + 1):
            header_value = self.index_sheet.cell(row=5, column=col_idx).value
            if header_value is not None and str(header_value).strip() != "":
                header_to_col[str(header_value)] = col_idx

        ds_col_idx = header_to_col.get("ПРОЦЕНТИ ЗА ДЕПОЗИТНИМИ СЕРТИФІКАТАМИ (ДС), тис. грн")
        ovdp_col_idx = header_to_col.get("ПРОЦЕНТИ ЗА ОВДП (ОВДП), тис.грн")
        pd_col_idx = header_to_col.get("ПРОЦЕНТНІ ДОХОДИ ВСЬОГО (ПД), тис. грн")
        ds_pd_col_idx = header_to_col.get("ДС/ПД, %")
        ds_share_col_idx = header_to_col.get("ДС до підсумку, %")
        ds_ovdp_pd_col_idx = header_to_col.get("(ДС+ОВДП)/ПД, %")
        ovdp_pd_col_idx = header_to_col.get("ОВДП/ПД, %")
        ovdp_share_col_idx = header_to_col.get("ОВДП до підсумку")
        finrez_col_idx = header_to_col.get("ФінРез (ОВДП-ДС)")

        if ds_col_idx is not None and pd_col_idx is not None and ds_pd_col_idx is not None:
            ds_col_letter, pd_col_letter, ds_pd_col_letter = get_column_letter(ds_col_idx), get_column_letter(
                pd_col_idx), get_column_letter(ds_pd_col_idx)
            for row_idx in range(6, self.data_last_row + 1):
                ds_pd_cell = self.index_sheet[f"{ds_pd_col_letter}{row_idx}"]
                ds_pd_cell.value = f"=IFERROR({ds_col_letter}{row_idx}/{pd_col_letter}{row_idx}*100,0)"
                ds_pd_cell.number_format = '0.0'

        if ds_col_idx is not None and ds_share_col_idx is not None:
            ds_col_letter, ds_share_col_letter = get_column_letter(ds_col_idx), get_column_letter(ds_share_col_idx)
            for row_idx in range(6, self.data_last_row + 1):
                ds_share_cell = self.index_sheet[f"{ds_share_col_letter}{row_idx}"]
                ds_share_cell.value = f"=IFERROR({ds_col_letter}{row_idx}/SUM({ds_col_letter}6:{ds_col_letter}{self.data_last_row})*100,0)"
                ds_share_cell.number_format = '0.0'

        if ds_col_idx is not None and ovdp_col_idx is not None and pd_col_idx is not None and ds_ovdp_pd_col_idx is not None:
            ds_col_letter, ovdp_col_letter, pd_col_letter, ds_ovdp_pd_col_letter = get_column_letter(
                ds_col_idx), get_column_letter(ovdp_col_idx), get_column_letter(pd_col_idx), get_column_letter(
                ds_ovdp_pd_col_idx)
            for row_idx in range(6, self.data_last_row + 1):
                ds_ovdp_pd_cell = self.index_sheet[f"{ds_ovdp_pd_col_letter}{row_idx}"]
                ds_ovdp_pd_cell.value = f"=IFERROR(({ds_col_letter}{row_idx}+{ovdp_col_letter}{row_idx})/{pd_col_letter}{row_idx}*100,0)"
                ds_ovdp_pd_cell.number_format = '0.0'

        if ovdp_col_idx is not None and pd_col_idx is not None and ovdp_pd_col_idx is not None:
            ovdp_col_letter, pd_col_letter, ovdp_pd_col_letter = get_column_letter(ovdp_col_idx), get_column_letter(
                pd_col_idx), get_column_letter(ovdp_pd_col_idx)
            for row_idx in range(6, self.data_last_row + 1):
                ovdp_pd_cell = self.index_sheet[f"{ovdp_pd_col_letter}{row_idx}"]
                ovdp_pd_cell.value = f"=IFERROR({ovdp_col_letter}{row_idx}/{pd_col_letter}{row_idx}*100,0)"
                ovdp_pd_cell.number_format = '0.0'

        if ovdp_col_idx is not None and ovdp_share_col_idx is not None:
            ovdp_col_letter, ovdp_share_col_letter = get_column_letter(ovdp_col_idx), get_column_letter(
                ovdp_share_col_idx)
            for row_idx in range(6, self.data_last_row + 1):
                ovdp_share_cell = self.index_sheet[f"{ovdp_share_col_letter}{row_idx}"]
                ovdp_share_cell.value = f"=IFERROR({ovdp_col_letter}{row_idx}/SUM({ovdp_col_letter}6:{ovdp_col_letter}{self.data_last_row})*100,0)"
                ovdp_share_cell.number_format = '0.0'

        if ds_col_idx is not None and ovdp_col_idx is not None and finrez_col_idx is not None:
            ds_col_letter, ovdp_col_letter, finrez_col_letter = get_column_letter(ds_col_idx), get_column_letter(
                ovdp_col_idx), get_column_letter(finrez_col_idx)
            for row_idx in range(6, self.data_last_row + 1):
                finrez_cell = self.index_sheet[f"{finrez_col_letter}{row_idx}"]
                finrez_cell.value = f"=IFERROR({ds_col_letter}{row_idx}+{ovdp_col_letter}{row_idx},0)"
                finrez_cell.number_format = '#,##0.00'

        for col_idx in range(6, self.index_sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_cell = self.index_sheet[f'{col_letter}5']
            header_value = header_cell.value

            if header_value is None or str(header_value).strip() == "": continue

            non_empty_header_cols.append(col_idx)
            sum_cell = self.index_sheet[f"{col_letter}{total_row}"]

            if ds_pd_col_idx is not None and col_idx == ds_pd_col_idx:
                if ds_col_idx is not None and pd_col_idx is not None:
                    sum_cell.value = f"=IFERROR({get_column_letter(ds_col_idx)}{total_row}/{get_column_letter(pd_col_idx)}{total_row}*100,0)"
                else:
                    sum_cell.value = "=0"
            elif ds_share_col_idx is not None and col_idx == ds_share_col_idx:
                if ds_col_idx is not None:
                    sum_cell.value = f"=IFERROR({get_column_letter(ds_col_idx)}{total_row}/SUM({get_column_letter(ds_col_idx)}6:{get_column_letter(ds_col_idx)}{self.data_last_row})*100,0)"
                else:
                    sum_cell.value = "=0"
            elif ds_ovdp_pd_col_idx is not None and col_idx == ds_ovdp_pd_col_idx:
                if ds_col_idx is not None and ovdp_col_idx is not None and pd_col_idx is not None:
                    sum_cell.value = f"=IFERROR(({get_column_letter(ds_col_idx)}{total_row}+{get_column_letter(ovdp_col_idx)}{total_row})/{get_column_letter(pd_col_idx)}{total_row}*100,0)"
                else:
                    sum_cell.value = "=0"
            elif ovdp_pd_col_idx is not None and col_idx == ovdp_pd_col_idx:
                if ovdp_col_idx is not None and pd_col_idx is not None:
                    sum_cell.value = f"=IFERROR({get_column_letter(ovdp_col_idx)}{total_row}/{get_column_letter(pd_col_idx)}{total_row}*100,0)"
                else:
                    sum_cell.value = "=0"
            elif ovdp_share_col_idx is not None and col_idx == ovdp_share_col_idx:
                if ovdp_col_idx is not None:
                    sum_cell.value = f"=IFERROR({get_column_letter(ovdp_col_idx)}{total_row}/SUM({get_column_letter(ovdp_col_idx)}6:{get_column_letter(ovdp_col_idx)}{self.data_last_row})*100,0)"
                else:
                    sum_cell.value = "=0"
            else:
                sum_cell.value = f"=IFERROR(SUM({col_letter}6:{col_letter}{self.data_last_row}),0)"

            sum_cell.font = Font(name='Arial', size=10, bold=True)
            sum_cell.alignment = Alignment(horizontal='center', vertical='center')
            sum_cell.number_format = '#,##0.00'

            results[str(header_value)] = {'column': col_letter, 'formula': sum_cell.value}

        if not non_empty_header_cols: return results

        last_formula_col = max(non_empty_header_cols)

        for col_idx in range(1, last_formula_col + 1):
            cell = self.index_sheet.cell(row=total_row, column=col_idx)
            cell.border = thin_border
            cell.fill = total_fill

        self.index_sheet.row_dimensions[total_row].height = 22

        # -------------------------------------------------------------
        # NEW LOGIC: Reading the Category directly from the newly created Column D
        # -------------------------------------------------------------
        state_rows, foreign_rows, private_rows, other_rows = [], [], [], []

        for row_idx in range(6, self.data_last_row + 1):
            cat_val = self.index_sheet.cell(row=row_idx, column=4).value
            if cat_val == "Державний":
                state_rows.append(row_idx)
            elif cat_val == "Іноземний капітал":
                foreign_rows.append(row_idx)
            elif cat_val == "Приватний капітал":
                private_rows.append(row_idx)
            else:
                other_rows.append(row_idx)

        state_row = total_row + 1
        foreign_row = total_row + 2
        private_row_total = total_row + 3
        other_row_total = total_row + 4

        category_rows = [
            (state_row, "Державні банки", state_rows),
            (foreign_row, "Банки з іноземним капіталом", foreign_rows),
            (private_row_total, "Банки з українським приватним капіталом", private_rows),
            (other_row_total, "Інші банки", other_rows)
        ]

        def _category_formula(rows_list, col_index):
            if not rows_list: return "=0"
            col_letter = get_column_letter(col_index)
            refs = ",".join([f"{col_letter}{r}" for r in sorted(rows_list)])
            return f"=IFERROR(SUM({refs}),0)"

        for out_row, category_name, source_rows in category_rows:
            self.index_sheet.cell(row=out_row, column=1).value = category_name
            self.index_sheet.cell(row=out_row, column=1).font = Font(name='Arial', size=10)
            self.index_sheet.cell(row=out_row, column=1).alignment = Alignment(horizontal='left', vertical='center')

            self.index_sheet.cell(row=out_row, column=4).value = f"{len(source_rows)} " + "банк(и/ів)"
            self.index_sheet.cell(row=out_row, column=4).font = Font(name='Arial', size=10, bold=True)
            self.index_sheet.cell(row=out_row, column=4).alignment = Alignment(horizontal='center', vertical='center')

            for col_idx in non_empty_header_cols:
                out_cell = self.index_sheet.cell(row=out_row, column=col_idx)
                if ds_pd_col_idx is not None and col_idx == ds_pd_col_idx:
                    if ds_col_idx is not None and pd_col_idx is not None:
                        out_cell.value = f"=IFERROR({get_column_letter(ds_col_idx)}{out_row}/{get_column_letter(pd_col_idx)}{out_row}*100,0)"
                    else:
                        out_cell.value = "=0"
                elif ds_share_col_idx is not None and col_idx == ds_share_col_idx:
                    if ds_col_idx is not None:
                        out_cell.value = f"=IFERROR({get_column_letter(ds_col_idx)}{out_row}/{get_column_letter(ds_col_idx)}{total_row}*100,0)"
                    else:
                        out_cell.value = "=0"
                elif ovdp_share_col_idx is not None and col_idx == ovdp_share_col_idx:
                    if ovdp_col_idx is not None:
                        out_cell.value = f"=IFERROR({get_column_letter(ovdp_col_idx)}{out_row}/{get_column_letter(ovdp_col_idx)}{total_row}*100,0)"
                    else:
                        out_cell.value = "=0"
                else:
                    out_cell.value = _category_formula(source_rows, col_idx)
                out_cell.number_format = '#,##0.00'
                out_cell.alignment = Alignment(horizontal='center', vertical='center')

            for col_idx in range(1, last_formula_col + 1):
                cell = self.index_sheet.cell(row=out_row, column=col_idx)
                cell.border = thin_border
                cell.fill = total_fill

            self.index_sheet.row_dimensions[out_row].height = 22

        print(f"        Totals rows filled out")
        print(f"        Creating ancillary tables")

        # -------------------------------------------------------------
        # STAGE 4 & 5: ANCILLARY TABLES AND CHARTS (UNTOUCHED)
        # -------------------------------------------------------------
        middle_col_name = 2
        middle_col_value = 3
        right_col_name = 8
        right_col_value = 9
        table_header_row = other_row_total + 2

        categories = [
            ("Державні банки", state_row),
            ("Банки з іноземним капіталом", foreign_row),
            ("Банки з українським приватним капіталом", private_row_total),
            ("Інші банки", other_row_total)
        ]

        self.index_sheet.cell(row=table_header_row, column=middle_col_name).value = "Категорія"
        self.index_sheet.cell(row=table_header_row, column=middle_col_value).value = "ДС/ПД, %"
        self.index_sheet.cell(row=table_header_row, column=right_col_name).value = "Категорія"
        self.index_sheet.cell(row=table_header_row, column=right_col_value).value = "ДС до підсумку, %"

        for idx, (category_label, category_row) in enumerate(categories, start=1):
            out_row = table_header_row + idx

            self.index_sheet.cell(row=out_row, column=middle_col_name).value = f"=A{category_row}"
            if ds_pd_col_idx is not None:
                ds_pd_letter = get_column_letter(ds_pd_col_idx)
                self.index_sheet.cell(row=out_row,
                                      column=middle_col_value).value = f"=IFERROR({ds_pd_letter}{category_row},0)"
            else:
                self.index_sheet.cell(row=out_row, column=middle_col_value).value = "=0"

            self.index_sheet.cell(row=out_row, column=right_col_name).value = f"=A{category_row}"
            if ds_share_col_idx is not None:
                ds_share_letter = get_column_letter(ds_share_col_idx)
                self.index_sheet.cell(row=out_row,
                                      column=right_col_value).value = f"=IFERROR({ds_share_letter}{category_row},0)"
            else:
                self.index_sheet.cell(row=out_row, column=right_col_value).value = "=0"

        table_cols = [middle_col_name, middle_col_value, right_col_name, right_col_value]
        for r in range(table_header_row, table_header_row + 5):
            for c in table_cols:
                cell = self.index_sheet.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if r == table_header_row:
                    cell.font = Font(name='Arial', size=10, bold=True)
                    cell.fill = total_fill
                else:
                    if c in [middle_col_value, right_col_value]:
                        cell.number_format = '0.0'

        self.index_sheet.column_dimensions[get_column_letter(middle_col_name)].width = 38
        self.index_sheet.column_dimensions[get_column_letter(middle_col_value)].width = 14
        self.index_sheet.column_dimensions[get_column_letter(right_col_name)].width = 38
        self.index_sheet.column_dimensions[get_column_letter(right_col_value)].width = 16

        print(f"        Both ancillary tables created")
        print(f"        Creating Index Sheet charts")

        middle_cats = Reference(self.index_sheet, min_col=middle_col_name, min_row=table_header_row + 1,
                                max_row=table_header_row + 4)
        middle_vals = Reference(self.index_sheet, min_col=middle_col_value, min_row=table_header_row + 1,
                                max_row=table_header_row + 4)

        bar = BarChart()
        bar.type = "col"
        bar.title = "Частка процентів по депсертифікатах у\nпроцентних доходах банків, %"
        bar.add_data(middle_vals, titles_from_data=False)
        bar.set_categories(middle_cats)
        bar.height = 13
        bar.width = 13
        bar.legend = None
        bar.x_axis.delete = False
        bar.y_axis.delete = False
        bar.x_axis.title = None
        bar.y_axis.title = None
        bar.dataLabels = DataLabelList()
        bar.dataLabels.showSerName = False
        bar.dataLabels.showCatName = False
        bar.dataLabels.showVal = True
        bar.dataLabels.txPr = _datalabel_text_style(size_pt=13, bold=True, color="FF0000")

        right_cats = Reference(self.index_sheet, min_col=right_col_name, min_row=table_header_row + 1,
                               max_row=table_header_row + 4)
        right_vals = Reference(self.index_sheet, min_col=right_col_value, min_row=table_header_row + 1,
                               max_row=table_header_row + 4)

        pie_right = PieChart()
        pie_right.title = "Проценти за депозитними сертифікатами\nу розрізі банків, %"
        pie_right.title.txPr = _datalabel_text_style(size_pt=13, bold=True, color="FF0000")
        pie_right.add_data(right_vals, titles_from_data=False)
        pie_right.set_categories(right_cats)
        pie_right.height = 14
        pie_right.width = 14
        pie_right.legend = None
        pie_right.dataLabels = DataLabelList()
        pie_right.dataLabels.showSerName = False
        pie_right.dataLabels.showCatName = True
        pie_right.dataLabels.showVal = False
        pie_right.dataLabels.showPercent = True
        pie_right.dataLabels.showLeaderLines = True
        pie_right.dataLabels.txPr = _datalabel_text_style(size_pt=11, bold=True, color="000000")

        chart_top_row = table_header_row + 6
        self.index_sheet.add_chart(bar, f"B{chart_top_row}")
        self.index_sheet.add_chart(pie_right, f"H{chart_top_row}")

        print(f"        Index Sheet charts created")
        return results