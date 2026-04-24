from .utils import *
from ..account_rows import BalanceSheet
import time
import openpyxl as xl
from multiprocessing import Pool, cpu_count

# --- 1. GLOBAL WORKER VARIABLES ---
# These variables exist in the isolated memory of each separate CPU core.
worker_wb = None
worker_file_path = None


def init_worker(file_path):
    """Runs exactly ONCE per CPU core when the Pool starts."""
    global worker_wb, worker_file_path
    worker_file_path = file_path

    print(f"        [Core Initializing] Loading workbook into memory...")
    # FATAL BUG FIX: Removed read_only=True!
    # Openpyxl crashes if you use read_only combined with self.sheet["A5"]
    worker_wb = xl.load_workbook(file_path, data_only=True)


def _process_single_sheet_worker(sheet_name):
    """Runs for every sheet. Uses the already-loaded workbook in memory."""
    global worker_wb, worker_file_path

    try:
        target_sheet = worker_wb[sheet_name]

        # Runs the lean 4-second __init__ using the pre-loaded sheet
        bs = BalanceSheet(parent_file=worker_file_path, sheet=target_sheet)

        return {
            "sheet_name": sheet_name,
            "indicator_frame": bs.indicator_frame,
            "pivot_db": bs.pivot_db,
            "start_row": bs.start_row,
            "start_col": bs.start_col
        }
    except Exception as e:
        # NO MORE SILENT CRASHES: If a worker dies, pass the error back!
        return {"ERROR": str(e), "sheet_name": sheet_name}


class IndicatorTablesCreator:
    def __init__(self, workbook, workbook_data, parent_file, progress_var):
        print(f"\n=== STAGE 2: Indicator Tables Creation ===")
        print(f"    Initialising an IndicatorTablesCreator object")
        self.workbook = workbook
        self.workbook_data = workbook_data
        self.parent_file = parent_file
        self.progress_var = progress_var
        self.delay = 0.01
        self.percentage = 55
        self.sheets_dict = {}
        self.create_sheets_dict()
        print(f"    STAGE 2 COMPLETE")

    def update_progress(self, value):
        current = self.progress_var.get()
        new = current + value / 100.0
        self.progress_var.set(new)
        time.sleep(self.delay)

    def create_sheets_dict(self):
        print(f"    Creating BalanceSheet objects:")

        # Filter valid sheets to process
        valid_sheets = [s.title for s in self.workbook.worksheets if code_bank(s.title) is not None]
        total = len(valid_sheets)

        if total == 0:
            return

        increment_percent = self.percentage / total

        print(f"    Firing up {cpu_count()} CPU cores to process {total} sheets...")

        # 2. FIRE UP THE POOL (Notice the initializer arguments!)
        with Pool(processes=cpu_count(), initializer=init_worker, initargs=(self.parent_file,)) as pool:
            # We just pass the string names of the sheets to the workers
            for i, result in enumerate(pool.imap_unordered(_process_single_sheet_worker, valid_sheets)):

                # Check if the worker experienced an error
                if "ERROR" in result:
                    print(f"        [!] ERROR IN SHEET {result['sheet_name']}: {result['ERROR']}")
                    continue

                sheet_name = result["sheet_name"]
                title = code_bank(sheet_name)

                print(f"        [{i + 1}/{total}] ✓ Math complete for: {sheet_name}")

                # Update GUI progress
                self.update_progress(increment_percent)

                # 3. WRITE SEQUENTIALLY TO MASTER WORKBOOK
                target_sheet = self.workbook[sheet_name]

                dummy_bs = BalanceSheet.__new__(BalanceSheet)
                dummy_bs.sheet = target_sheet
                dummy_bs.indicator_frame = result["indicator_frame"]

                dummy_bs.insert_frame(
                    parent_file=self.parent_file,
                    start_row=result["start_row"],
                    start_col=result["start_col"]
                )

                dummy_bs.pivot_db = result["pivot_db"]
                self.sheets_dict[title] = dummy_bs

        print(f"    ✓ Created dictionary and wrote {len(self.sheets_dict)} indicator tables")