"""
Combined Bank Data Processing Script


This script combines the functionality of:
1. create_sheet_index.py - Creates and formats the Index sheet
2. Danya's_code.py - Fills the Index sheet with formulas from bank sheets

Process:
1. Create Index sheet with headers and formatting
2. Add hyperlinks to all bank sheets
3. Fill formulas linking to indicators in bank sheets
4. Calculate column sums
5. Save the file
"""

import openpyxl as xl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
# from openpyxl.xml.functions import fromstring
from xml.etree.ElementTree import fromstring
import sys
import io
import re
from pathlib import Path
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# from ..account_rows import BalanceSheet

# Ensure UTF-8 encoding for Cyrillic characters
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
if sys.stderr.encoding != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


def _datalabel_text_style(size_pt=12, bold=True, color="000000"):
    size_100 = int(size_pt * 100)
    bold_flag = "1" if bold else "0"
    xml = f'''
<a:txPr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
    <a:bodyPr/>
    <a:lstStyle/>
    <a:p>
        <a:pPr>
            <a:defRPr sz="{size_100}" b="{bold_flag}">
                <a:solidFill><a:srgbClr val="{color}"/></a:solidFill>
            </a:defRPr>
        </a:pPr>
        <a:endParaRPr lang="uk-UA"/>
    </a:p>
</a:txPr>
'''
    return RichText.from_tree(fromstring(xml))


def make_formula(sheet_name, column, row):
    """Create Excel formula reference to another sheet."""
    return f"='{sheet_name}'!${column}${row}"


def safe_set_cell_value(sheet, row, col, value):
    """Safely set cell value, handling merged cells"""
    target_cell = sheet.cell(row=row, column=col)

    # Check if it's a merged cell
    if isinstance(target_cell, MergedCell):
        # Find the merged range and unmerge it temporarily
        for merged_range in list(sheet.merged_cells.ranges):
            if target_cell.coordinate in merged_range:
                sheet.unmerge_cells(str(merged_range))
                # Now write to the cell
                sheet.cell(row=row, column=col).value = value
                return

    # If not merged, just set the value normally
    target_cell.value = value


def code_bank(name):
    """Extract bank code from sheet name."""
    match = re.match(r"\s*(\d+)", name)
    return (match.group(1), str(name)) if match else None


def is_valid_sheet_name(sheet_name):
    """Check if a sheet name matches the required pattern (starts with a number)."""
    pattern = r'^\s*\d+'
    return re.match(pattern, sheet_name) is not None

def _normalize_name(name):
    if name is None:
        return ""
    return re.sub(r"\s+", " ", str(name)).strip().lower()

STATE_BANKS = {
    _normalize_name('АТ "Укрексімбанк"'),
    _normalize_name('АТ "ОЩАДБАНК"'),
    _normalize_name('АТ КБ "ПриватБанк"'),
    _normalize_name('АТ "СЕНС БАНК"'),
    _normalize_name('АБ "УКРГАЗБАНК"'),
    _normalize_name('АТ "ПЕРШИЙ ІНВЕСТИЦІЙНИЙ БАНК"'),
    _normalize_name('АТ "МОТОР-БАНК"')
}

FOREIGN_BANKS = {
    _normalize_name('АТ "Райффайзен Банк"'),
    _normalize_name('АТ АКБ "Львів" '),
    _normalize_name('АТ "СКАЙ БАНК" '),
    _normalize_name('АТ "БТА БАНК" '),
    _normalize_name('АТ "УКРСИББАНК" '),
    _normalize_name('АТ "Ідея Банк" '),
    _normalize_name('АТ "ПРАВЕКС БАНК" '),
    _normalize_name('АТ "КРЕДІ АГРІКОЛЬ БАНК" '),
    _normalize_name('АТ "ЮНЕКС БАНК" '),
    _normalize_name('АТ "ПІРЕУС БАНК МКБ" '),
    _normalize_name('АТ "ІНГ Банк Україна" '),
    _normalize_name('АТ "ОТП БАНК" '),
    _normalize_name('АТ "СІТІБАНК" '),
    _normalize_name('АТ "ПРОКРЕДИТ БАНК" '),
    _normalize_name('АТ "УБРР"'),
    _normalize_name('АТ "НЕКСЕНТ БАНК" '),
    _normalize_name('АТ "КРЕДИТВЕСТ БАНК"'),
    _normalize_name('АТ "АГРОПРОСПЕРІС БАНК" '),
    _normalize_name('АТ "Дойче Банк ДБУ"  '),
    _normalize_name('АТ"СЕБ КОРПОРАТИВНИЙ БАНК"'),
    _normalize_name('АТ "КРЕДОБАНК"')
}